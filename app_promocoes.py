import streamlit as st
import requests
import pandas as pd
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

DEGUST_API_BASE = "https://lx-degust-api-integracao-prd.azurewebsites.net"

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io

# Configuração da página
st.set_page_config(
    page_title="Dashboard de Promoções",
    page_icon="🎉",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para melhorar a aparência
st.markdown("""
    <style>
    .main {
        padding: 0rem 1rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        font-weight: bold;
        border-radius: 5px;
        border: none;
        padding: 10px;
    }
    .stButton>button:hover {
        background-color: #45a049;
    }
    h1 {
        color: #2c3e50;
        text-align: center;
    }
    .dataframe {
        font-size: 12px;
    }
    </style>
""", unsafe_allow_html=True)

# Configurações das marcas
MARCAS_CONFIG = {
    "Promoções Bendito": {
        "codfranqueador": 3082,
        "cor": "#FF6B6B"
    },
    "Promoções Espetto": {
        "codfranqueador": 3078,
        "cor": "#4ECDC4"
    },
    "Promoções Mané": {
        "codfranqueador": 1428,
        "cor": "#95E1D3"
    },
    "Promoções Buteco Seu Rufino": {
        "codfranqueador": 3081,
        "cor": "#FFA500"
    }
}

CREDENCIAIS = {
    "usuario": "06266555794",
    "senha": "250913"
}

def autenticar(codfranqueador):
    """Realiza autenticação na API do Degust"""
    url_auth = "https://lx-degust-api-integracao-prd.azurewebsites.net/api/usuario/autenticar"
    
    credenciais = {
        "usuario": CREDENCIAIS["usuario"],
        "senha": CREDENCIAIS["senha"],
        "codigoFranqueador": codfranqueador
    }
    
    try:
        response = requests.post(url_auth, json=credenciais, timeout=10)
        if response.status_code == 200:
            token = response.json()["acesso"]["token"]
            return token
        else:
            st.error(f"❌ Erro ao autenticar: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"❌ Erro de conexão: {str(e)}")
        return None

def _loja_degust_ativa(loja):
    """
    Fallback quando GET /api/loja/loja nao retorna dadosGerais.ativo de forma conclusiva.
    """
    for key in ("situacaoLoja", "situacao"):
        val = loja.get(key)
        if val is None:
            continue
        s = str(val).strip().upper()
        if s and "INATIV" in s:
            return False
    return True


def _interpretar_campo_ativo_cadastro(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if not s:
        return None
    if s in ("S", "SIM", "1", "T", "TRUE", "Y", "YES", "ATIVO", "ATIVA"):
        return True
    if s in ("N", "NAO", "NÃO", "0", "F", "FALSE", "INATIVO", "INATIVA"):
        return False
    if "INATIV" in s:
        return False
    if "ATIV" in s:
        return True
    return None


def _consultar_ativo_cadastro_loja(cliente_http, token, codfranqueador, codigo_loja):
    url = f"{DEGUST_API_BASE}/api/loja/loja"
    params = {"CodigoFranqueador": int(codfranqueador), "CodigoLoja": int(codigo_loja)}
    headers = {"Authorization": f"Bearer {token}"}
    try:
        r = cliente_http.get(url, params=params, headers=headers, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        if not isinstance(data, dict):
            return None
        dg = data.get("dadosGerais") or {}
        if not isinstance(dg, dict):
            return None
        return _interpretar_campo_ativo_cadastro(dg.get("ativo"))
    except Exception:
        return None


def _manter_loja_apos_consulta_cadastro(loja, ativo_cadastro):
    if ativo_cadastro is False:
        return False
    if ativo_cadastro is True:
        return True
    return _loja_degust_ativa(loja)


def _filtrar_lojas_por_cadastro_degust(lojas, token, codfranqueador, max_workers=8):
    if not lojas:
        return []

    def processar(loja, http):
        cod = loja.get("codigoLoja")
        try:
            c = int(cod) if cod is not None else None
        except (TypeError, ValueError):
            c = None
        if c is None:
            return _manter_loja_apos_consulta_cadastro(loja, None)
        at = _consultar_ativo_cadastro_loja(http, token, codfranqueador, c)
        return _manter_loja_apos_consulta_cadastro(loja, at)

    if len(lojas) == 1:
        loja = lojas[0]
        return [loja] if processar(loja, requests) else []

    thread_local = threading.local()

    def http_por_thread():
        if not hasattr(thread_local, "session"):
            thread_local.session = requests.Session()
        return thread_local.session

    workers = max(1, min(max_workers, len(lojas)))

    def work(loja):
        return loja, processar(loja, http_por_thread())

    with ThreadPoolExecutor(max_workers=workers) as executor:
        pairs = list(executor.map(work, lojas))
    return [loja for loja, ok in pairs if ok]


def obter_lojas(token, codfranqueador):
    """Obtém lista de lojas da franquia com dados completos"""
    url_lojas = f"{DEGUST_API_BASE}/api/loja/listarLojasFranquia?codigoFranquia={codfranqueador}"
    headers = {"Authorization": f"Bearer {token}"}
    
    try:
        response = requests.get(url_lojas, headers=headers, timeout=10)
        if response.status_code == 200:
            lojas = response.json()
            if not isinstance(lojas, list):
                return []
            return _filtrar_lojas_por_cadastro_degust(lojas, token, codfranqueador)
        else:
            st.error(f"❌ Erro ao buscar lojas: {response.status_code}")
            return []
    except Exception as e:
        st.error(f"❌ Erro de conexão: {str(e)}")
        return []

def consultar_promocoes(token, codfranqueador, lojas_completas, marca):
    """Consulta promoções de todas as lojas"""
    url_promocoes = "https://lx-degust-api-integracao-prd.azurewebsites.net/api/produto/consultar-promocoes"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    # Criar dicionário de mapeamento código -> nome da loja
    mapeamento_lojas = {}
    for loja in lojas_completas:
        if "codigoLoja" in loja and "nomeLoja" in loja:
            mapeamento_lojas[loja["codigoLoja"]] = loja["nomeLoja"]
    
    dados_todas_lojas = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_lojas = len(lojas_completas)
    
    for idx, loja in enumerate(lojas_completas):
        codigo_loja = loja["codigoLoja"]
        nome_loja = loja.get("nomeLoja", "N/A")
        
        status_text.text(f"🔄 Carregando promoções da loja {idx + 1}/{total_lojas}: {nome_loja}")
        
        body = {
            "codigoFranquia": codfranqueador,
            "codigoLoja": codigo_loja
        }
        
        try:
            response = requests.post(url_promocoes, json=body, headers=headers, timeout=10)
            
            if response.status_code == 200:
                dados = response.json()
                if dados:
                    for item in dados:
                        item["codigoLoja"] = codigo_loja
                        item["nomeLoja"] = nome_loja  # Adicionar nome da loja
                        item["marca"] = marca
                    dados_todas_lojas.extend(dados)
        except Exception as e:
            st.warning(f"⚠️ Erro ao buscar promoções da loja {nome_loja} ({codigo_loja}): {str(e)}")
        
        progress_bar.progress((idx + 1) / total_lojas)
    
    progress_bar.empty()
    status_text.empty()
    
    return dados_todas_lojas

def consultar_produtos_grupo_venda_orientada(token, codfranqueador, lojas_completas, marca, nome_grupo="Promoção"):
    """Consulta produtos por grupo de venda orientada de todas as lojas usando autenticação da API
    
    IMPORTANTE: Utiliza as MESMAS credenciais e token da API de promoções.
    O token é obtido através da função autenticar() que usa CREDENCIAIS["usuario"] e CREDENCIAIS["senha"].
    """
    url = "https://lx-degust-api-integracao-prd.azurewebsites.net/api/venda-orientada/consultar-produto-por-grupo-venda-orientada"
    
    # Usar o MESMO token e formato de autenticação da API de promoções
    # O token é obtido da mesma função autenticar() que usa as mesmas credenciais
    headers = {
        "Authorization": f"Bearer {token}",  # Mesmo token usado na API de promoções
        "Content-Type": "application/json"
    }
    
    dados_todas_lojas = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_lojas = len(lojas_completas)
    
    for idx, loja in enumerate(lojas_completas):
        codigo_loja = loja["codigoLoja"]
        nome_loja = loja.get("nomeLoja", "N/A")
        
        status_text.text(f"🔄 Carregando produtos do grupo '{nome_grupo}' da loja {idx + 1}/{total_lojas}: {nome_loja}")
        
        # Body com os mesmos campos usados na API de promoções
        body = {
            "codigoFranquia": codfranqueador,
            "codigoLoja": codigo_loja,
            "nomeGrupoVendaOrientada": nome_grupo
        }
        
        try:
            # Usar o MESMO token de autenticação obtido da função autenticar()
            # que utiliza as MESMAS credenciais (CREDENCIAIS["usuario"] e CREDENCIAIS["senha"])
            # usadas na API de promoções
            response = requests.post(url, json=body, headers=headers, timeout=10)
            
            # Verificar se houve erro de autenticação
            if response.status_code == 401:
                st.warning(f"⚠️ Erro de autenticação ao consultar grupo '{nome_grupo}' da loja {nome_loja}. Token pode ter expirado.")
                continue
            elif response.status_code == 403:
                st.warning(f"⚠️ Acesso negado ao consultar grupo '{nome_grupo}' da loja {nome_loja}.")
                continue
            
            if response.status_code == 200:
                dados = response.json()
                if dados and isinstance(dados, list) and len(dados) > 0:
                    for item in dados:
                        item["codigoLoja"] = codigo_loja
                        item["nomeLoja"] = nome_loja
                        item["marca"] = marca
                        item["grupoVendaOrientada"] = nome_grupo
                        # Criar um nome de promoção baseado no grupo e loja
                        item["nomePromocao"] = f"PROMOÇÕES - {nome_loja.upper()}"
                        item["nomeGrupo"] = "PROMOÇÕES DA UNIDADE"  # Nome do grupo conforme layout da imagem
                        item["promocaoAtiva"] = "Sim"  # Assumir ativo se retornou dados
                    dados_todas_lojas.extend(dados)
            elif response.status_code != 200:
                # Log silencioso para outros erros (grupo pode não existir)
                pass
        except Exception as e:
            # Não exibir warning para grupos que não existem (é esperado)
            # Mas logar erros de conexão importantes
            if "timeout" in str(e).lower() or "connection" in str(e).lower():
                st.warning(f"⚠️ Erro de conexão ao consultar grupo '{nome_grupo}' da loja {nome_loja}: {str(e)}")
        
        progress_bar.progress((idx + 1) / total_lojas)
    
    progress_bar.empty()
    status_text.empty()
    
    return dados_todas_lojas

def criar_excel_formatado(df):
    """Cria um arquivo Excel formatado a partir de um DataFrame"""
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Promoções"
    
    # Adicionar dados do DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar larguras das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Limitar largura máxima e mínima
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Formatar células de dados (alinhamento)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def analisar_promocoes_por_cobertura(df_marca):
    """Analisa promoções por cobertura de lojas"""
    if df_marca.empty or "nomePromocao" not in df_marca.columns or "codigoLoja" not in df_marca.columns:
        return {}
    
    # Contar quantas lojas cada promoção aparece
    promocao_por_loja = df_marca.groupby("nomePromocao")["codigoLoja"].nunique()
    total_lojas = df_marca["codigoLoja"].nunique()
    
    # Calcular percentual de cobertura
    promocao_cobertura = promocao_por_loja / total_lojas * 100
    
    # Categorizar promoções
    promocoes_100 = promocao_por_loja[promocao_por_loja == total_lojas]
    promocoes_80_plus = promocao_por_loja[promocao_por_loja >= total_lojas * 0.8]
    promocoes_50_plus = promocao_por_loja[promocao_por_loja >= total_lojas * 0.5]
    
    # Top 10 promoções mais comuns
    top_10 = promocao_por_loja.nlargest(10)
    
    # Criar DataFrame com análise
    df_analise = pd.DataFrame({
        'Promoção': promocao_por_loja.index,
        'Lojas': promocao_por_loja.values,
        'Cobertura (%)': promocao_cobertura.values.round(1)
    }).sort_values('Cobertura (%)', ascending=False)
    
    # Adicionar status das promoções
    if "promocaoAtiva" in df_marca.columns:
        status_promocoes = df_marca.groupby("nomePromocao")["promocaoAtiva"].first()
        df_analise['Status'] = df_analise['Promoção'].map(status_promocoes).fillna('N/A')
    
    return {
        'total_lojas': total_lojas,
        'total_promocoes': len(promocao_por_loja),
        'promocoes_100': promocoes_100,
        'promocoes_80_plus': promocoes_80_plus,
        'promocoes_50_plus': promocoes_50_plus,
        'top_10': top_10,
        'df_analise': df_analise,
        'media_cobertura': promocao_cobertura.mean(),
        'mediana_cobertura': promocao_cobertura.median()
    }

@st.cache_data(ttl=300)  # Cache por 5 minutos
def carregar_dados_marca(marca):
    """Carrega dados de uma marca específica, incluindo promoções e produtos do grupo de venda orientada"""
    config = MARCAS_CONFIG[marca]
    codfranqueador = config["codfranqueador"]
    
    # Validar código do franqueador
    if not codfranqueador or codfranqueador == 0:
        st.warning(f"⚠️ Código do franqueador não configurado para {marca}. Por favor, atualize o código em MARCAS_CONFIG.")
        return pd.DataFrame()
    
    # Autenticar
    token = autenticar(codfranqueador)
    if not token:
        return pd.DataFrame()
    
    # Obter lojas com dados completos
    lojas_completas = obter_lojas(token, codfranqueador)
    if not lojas_completas:
        return pd.DataFrame()
    
    # Consultar promoções normais
    dados = consultar_promocoes(token, codfranqueador, lojas_completas, marca)
    
    # Consultar produtos do grupo de venda orientada "Promoção"
    dados_grupo_venda = consultar_produtos_grupo_venda_orientada(
        token, 
        codfranqueador, 
        lojas_completas, 
        marca, 
        nome_grupo="Promoção"
    )
    
    # ALTERNATIVA: Se a API não retornou dados, extrair produtos das promoções normais
    # que tenham nomeGrupo="PROMOÇÕES DA UNIDADE" e nomePromocao contendo "PROMOCOES"
    # (esses são os produtos do grupo "Promoção")
    if not dados_grupo_venda and dados:
        dados_grupo_venda_alternativa = []
        for item in dados:
            # Verificar se é produto do grupo de venda orientada "Promoção"
            nome_grupo = str(item.get('nomeGrupo', '')).strip().upper()
            nome_promocao = str(item.get('nomePromocao', '')).strip().upper()
            
            # Produtos com nomeGrupo="PROMOÇÕES DA UNIDADE" e nomePromocao contendo "PROMOCOES"
            # são do grupo "Promoção" - independente do gvoCodigo
            if "PROMOÇÕES DA UNIDADE" in nome_grupo or "PROMOCOES DA UNIDADE" in nome_grupo:
                # Verificar também se o nome da promoção contém "PROMOCOES" ou "PROMOÇÕES"
                if "PROMOCOES" in nome_promocao or "PROMOÇÕES" in nome_promocao:
                    # Criar cópia do item e marcar como grupo de venda orientada
                    item_grupo = item.copy()
                    item_grupo["grupoVendaOrientada"] = "PROMOCAO"
                    dados_grupo_venda_alternativa.append(item_grupo)
        
        if dados_grupo_venda_alternativa:
            dados_grupo_venda = dados_grupo_venda_alternativa
    
    # Combinar dados de promoções normais e grupo de venda orientada
    if dados_grupo_venda:
        dados.extend(dados_grupo_venda)
    
    if dados:
        df = pd.DataFrame(dados)
        return df
    else:
        return pd.DataFrame()

def main():
    # Título
    st.title("🎉 Dashboard de Promoções das Marcas TT")
    st.markdown("---")
    
    # Sidebar com filtros e controles
    with st.sidebar:
        st.header("⚙️ Controles")
        
        # Botão de atualizar
        if st.button("🔄 Atualizar Dados", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        
        st.markdown("---")
        
        # Filtro de marcas
        st.header("🏪 Filtrar por Marca")
        
        filtro_todas = st.checkbox("Todas as Marcas", value=True)
        
        if filtro_todas:
            marcas_selecionadas = list(MARCAS_CONFIG.keys())
        else:
            marcas_selecionadas = []
            for marca in MARCAS_CONFIG.keys():
                if st.checkbox(marca, value=False):
                    marcas_selecionadas.append(marca)
        
        st.markdown("---")
        st.markdown(f"**📅 Última atualização:**  \n{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    
    # Carregar e exibir dados
    if not marcas_selecionadas:
        st.warning("⚠️ Selecione pelo menos uma marca para visualizar as promoções.")
        return
    
    # Carregar dados de todas as marcas selecionadas
    todos_dados = []
    
    for marca in marcas_selecionadas:
        with st.spinner(f"Carregando dados de {marca}..."):
            df_marca = carregar_dados_marca(marca)
            if not df_marca.empty:
                todos_dados.append(df_marca)
    
    if not todos_dados:
        st.error("❌ Nenhum dado encontrado para as marcas selecionadas.")
        return
    
    # Combinar todos os dataframes
    df_final = pd.concat(todos_dados, ignore_index=True)
    
    # Botão de download geral (se todas as marcas foram selecionadas)
    if filtro_todas and len(marcas_selecionadas) == len(MARCAS_CONFIG):
        st.markdown("### 📥 Download Geral")
        excel_geral = criar_excel_formatado(df_final)
        st.download_button(
            label="⬇️ Download Todas as Marcas - Todas as Lojas (Excel)",
            data=excel_geral,
            file_name=f"todas_marcas_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_geral"
        )
        st.markdown("---")
    
    # Métricas gerais
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📦 Total de Promoções", len(df_final))
    
    with col2:
        lojas_unicas = df_final["codigoLoja"].nunique() if "codigoLoja" in df_final.columns else 0
        st.metric("🏪 Lojas", lojas_unicas)
    
    with col3:
        marcas_unicas = df_final["marca"].nunique() if "marca" in df_final.columns else 0
        st.metric("🏷️ Marcas", marcas_unicas)
    
    with col4:
        produtos_unicos = df_final["codigoProduto"].nunique() if "codigoProduto" in df_final.columns else 0
        st.metric("🎯 Produtos Únicos", produtos_unicos)
    
    st.markdown("---")
    
    # Exibir tabela por marca
    for marca in marcas_selecionadas:
        df_marca = df_final[df_final["marca"] == marca] if "marca" in df_final.columns else pd.DataFrame()
        
        if not df_marca.empty:
            cor = MARCAS_CONFIG[marca]["cor"]
            
            st.markdown(f"### <span style='color: {cor}'>{marca}</span>", unsafe_allow_html=True)
            
            # Informações da tabela
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📦 Total de Promoções", len(df_marca))
            with col2:
                st.metric("📊 Total de Colunas", len(df_marca.columns))
            with col3:
                st.metric("🏪 Lojas Únicas", df_marca["codigoLoja"].nunique() if "codigoLoja" in df_marca.columns else 0)
            with col4:
                if "nomeLoja" in df_marca.columns:
                    lojas_nomes = df_marca["nomeLoja"].nunique()
                    col4_1, col4_2 = st.columns([3, 1])
                    with col4_1:
                        st.metric("🏷️ Lojas com Nome", lojas_nomes)
                    with col4_2:
                        st.markdown("**Promoções de Rede**")
                        if st.button("📊", help="Ver promoções por cobertura de lojas", key=f"btn_promocoes_{marca}", use_container_width=True):
                            st.session_state[f'show_modal_{marca}'] = True
                else:
                    st.metric("🏷️ Lojas com Nome", 0)
            
            # Análise de Promoções de Rede
            if f'show_modal_{marca}' in st.session_state and st.session_state[f'show_modal_{marca}']:
                with st.expander(f"📊 Promoções de Rede - {marca}", expanded=True):
                    st.markdown(f"### 📊 Análise de Promoções por Cobertura - {marca}")
                    
                    # Realizar análise
                    analise = analisar_promocoes_por_cobertura(df_marca)
                    
                    if analise:
                        # Métricas principais
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("🏪 Total de Lojas", analise['total_lojas'])
                        with col2:
                            st.metric("📦 Total de Promoções", analise['total_promocoes'])
                        with col3:
                            st.metric("📊 Média de Cobertura", f"{analise['media_cobertura']:.1f}%")
                        with col4:
                            st.metric("📈 Mediana de Cobertura", f"{analise['mediana_cobertura']:.1f}%")
                        
                        st.markdown("---")
                        
                        # Promoções por categoria de cobertura
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.markdown("#### 🎯 Promoções em TODAS as Lojas (100%)")
                            if len(analise['promocoes_100']) > 0:
                                for promocao, count in analise['promocoes_100'].items():
                                    st.markdown(f"✅ **{promocao}** ({count} lojas)")
                            else:
                                st.info("Nenhuma promoção aparece em todas as lojas")
                        
                        with col2:
                            st.markdown("#### 🚀 Promoções em 80%+ das Lojas")
                            promocoes_80 = analise['promocoes_80_plus']
                            if len(promocoes_80) > 0:
                                for promocao, count in promocoes_80.items():
                                    percentual = (count / analise['total_lojas']) * 100
                                    st.markdown(f"🟢 **{promocao}** ({count}/{analise['total_lojas']} - {percentual:.1f}%)")
                            else:
                                st.info("Nenhuma promoção com 80%+ de cobertura")
                        
                        with col3:
                            st.markdown("#### 📈 Promoções em 50%+ das Lojas")
                            promocoes_50 = analise['promocoes_50_plus']
                            if len(promocoes_50) > 0:
                                for promocao, count in promocoes_50.items():
                                    percentual = (count / analise['total_lojas']) * 100
                                    st.markdown(f"🟡 **{promocao}** ({count}/{analise['total_lojas']} - {percentual:.1f}%)")
                            else:
                                st.info("Nenhuma promoção com 50%+ de cobertura")
                        
                        st.markdown("---")
                        
                        # Top 10 promoções mais comuns
                        st.markdown("#### 🏆 Top 10 Promoções Mais Comuns")
                        if len(analise['top_10']) > 0:
                            for i, (promocao, count) in enumerate(analise['top_10'].items(), 1):
                                percentual = (count / analise['total_lojas']) * 100
                                st.markdown(f"{i:2d}. **{promocao}** - {count}/{analise['total_lojas']} lojas ({percentual:.1f}%)")
                        
                        st.markdown("---")
                        
                        # Tabela completa
                        st.markdown("#### 📋 Tabela Completa de Promoções")
                        st.dataframe(
                            analise['df_analise'],
                            use_container_width=True,
                            height=400
                        )
                        
                        # Botão para fechar análise
                        if st.button("❌ Fechar Análise", key=f"close_modal_{marca}"):
                            st.session_state[f'show_modal_{marca}'] = False
                            st.rerun()
                    else:
                        st.error("❌ Não foi possível analisar as promoções. Verifique se os dados estão corretos.")
                        if st.button("❌ Fechar Análise", key=f"close_modal_{marca}"):
                            st.session_state[f'show_modal_{marca}'] = False
                            st.rerun()
            
            # Mostrar lojas disponíveis com nomes e campo de busca
            if "nomeLoja" in df_marca.columns and "codigoLoja" in df_marca.columns:
                lojas_info = df_marca[["codigoLoja", "nomeLoja"]].drop_duplicates().sort_values("codigoLoja")
                
                # Campo de busca para filtrar lojas
                st.markdown("**🔍 Buscar Lojas:**")
                busca_loja = st.text_input(
                    "Digite o nome ou código da loja para filtrar:",
                    key=f"busca_loja_{marca}",
                    placeholder="Ex: DOWNTOWN ou 1"
                )
                
                # Filtrar lojas baseado na busca
                if busca_loja:
                    busca_lower = busca_loja.lower()
                    lojas_filtradas = lojas_info[
                        lojas_info["nomeLoja"].str.lower().str.contains(busca_lower, na=False) |
                        lojas_info["codigoLoja"].astype(str).str.contains(busca_lower, na=False)
                    ]
                else:
                    lojas_filtradas = lojas_info
                
                # Criar lista de lojas formatadas para o multiselect
                lojas_formatadas = []
                for _, loja in lojas_filtradas.iterrows():
                    lojas_formatadas.append(f"{loja['codigoLoja']} - {loja['nomeLoja']}")
                
                # Seletor de lojas para download
                st.markdown("**📥 Selecionar Lojas para Download:**")
                lojas_selecionadas_download = st.multiselect(
                    "Escolha as lojas para download (deixe vazio para todas as lojas da marca):",
                    options=lojas_formatadas,
                    default=[],
                    key=f"lojas_download_{marca}",
                    help="Se nenhuma loja for selecionada, o download incluirá todas as lojas da marca"
                )
                
                # Mostrar lojas disponíveis (todas, não apenas filtradas)
                st.markdown("**🏪 Lojas disponíveis:**")
                lojas_texto = []
                for _, loja in lojas_info.iterrows():
                    lojas_texto.append(f"{loja['codigoLoja']} - {loja['nomeLoja']}")
                st.markdown(", ".join(lojas_texto))
            else:
                lojas_selecionadas_download = []
            
            # Mostrar TODAS as colunas disponíveis
            st.markdown("**📋 Colunas disponíveis:** " + ", ".join(df_marca.columns.tolist()))
            
            # Opção para filtrar colunas (opcional)
            with st.expander("🔧 Opções de Exibição", expanded=False):
                colunas_selecionadas = st.multiselect(
                    "Selecione as colunas para exibir (deixe vazio para mostrar todas):",
                    options=df_marca.columns.tolist(),
                    default=df_marca.columns.tolist(),
                    key=f"colunas_{marca}"
                )
            
            # Reordenar colunas para colocar nomeLoja antes de nomePromocao
            colunas_ordenadas = []
            
            # Primeiro, adicionar colunas importantes na ordem desejada
            colunas_importantes = ["codigoLoja", "nomeLoja"]
            for col in colunas_importantes:
                if col in df_marca.columns:
                    colunas_ordenadas.append(col)
            
            # Depois, adicionar nomePromocao se existir
            if "nomePromocao" in df_marca.columns:
                colunas_ordenadas.append("nomePromocao")
            
            # Adicionar as demais colunas (exceto as já adicionadas)
            for col in df_marca.columns:
                if col not in colunas_ordenadas:
                    colunas_ordenadas.append(col)
            
            # Aplicar filtro de colunas selecionadas se houver
            if colunas_selecionadas:
                colunas_ordenadas = [col for col in colunas_ordenadas if col in colunas_selecionadas]
            
            # Reordenar o DataFrame
            df_exibir = df_marca[colunas_ordenadas]
            
            st.dataframe(
                df_exibir,
                use_container_width=True,
                height=600  # Aumentei a altura para acomodar mais dados
            )
            
            # Preparar dados para download baseado na seleção
            df_download = df_exibir.copy()
            
            # Se lojas foram selecionadas, filtrar apenas essas lojas
            if "nomeLoja" in df_marca.columns and "codigoLoja" in df_marca.columns and lojas_selecionadas_download:
                # Extrair códigos das lojas selecionadas
                codigos_selecionados = []
                for loja_str in lojas_selecionadas_download:
                    codigo = loja_str.split(" - ")[0]
                    try:
                        codigos_selecionados.append(int(codigo))
                    except:
                        codigos_selecionados.append(codigo)
                
                # Filtrar DataFrame
                df_download = df_download[df_download["codigoLoja"].isin(codigos_selecionados)]
                
                # Label do botão de download
                if len(lojas_selecionadas_download) == 1:
                    label_download = f"⬇️ Download {marca} - Loja Selecionada (Excel)"
                else:
                    label_download = f"⬇️ Download {marca} - {len(lojas_selecionadas_download)} Lojas Selecionadas (Excel)"
            else:
                # Todas as lojas da marca
                label_download = f"⬇️ Download {marca} - Todas as Lojas (Excel)"
            
            # Botão de download com colunas ordenadas (Excel formatado)
            excel_file = criar_excel_formatado(df_download)
            st.download_button(
                label=label_download,
                data=excel_file,
                file_name=f"{marca.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{marca}"
            )
            
            st.markdown("---")

if __name__ == "__main__":
    main()

