import streamlit as st
import requests
import pandas as pd
from collections import defaultdict
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import json
import html

from hmg_promocoes_unidade import (
    BASE_PRD,
    COLUNAS_RETAGUARDA,
)

COLUNAS_VO_PROMOCAO = COLUNAS_RETAGUARDA
COLUNAS_VO_PROMOCAO_OCULTAS = {
    "Grupo",
    "Descrição do Grupo",
    "Descrição Monitor",
    "domingo",
    "segunda",
    "terca",
    "quarta",
    "quinta",
    "sexta",
    "sabado",
    "restricaoHorario",
    "valorMix",
    "valorPromocionalMix",
}
COLUNAS_VO_PROMOCAO_EXIBICAO = [
    c for c in COLUNAS_VO_PROMOCAO if c not in COLUNAS_VO_PROMOCAO_OCULTAS
]

# Ambiente fixo: somente API de produção (PRD)
DEGUST_API_BASE = BASE_PRD
if "hmg" in DEGUST_API_BASE.lower():
    raise RuntimeError("Este dashboard aceita apenas a API PRD de integração Degust.")

# Configuração da página
st.set_page_config(
    page_title="Dashboard de Promoções",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para melhorar a aparência
st.markdown("""
    <style>
    .main {
        padding: 0rem 1rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        font-weight: bold;
        border-radius: 5px;
        border: none;
        padding: 10px;
    }
    .stButton>button:hover {
        background-color: #45a049;
    }
    h1 {
        color: #2c3e50;
        text-align: center;
    }
    .dataframe {
        font-size: 12px;
    }
    
    /* Estilos para layout hierárquico */
    .promocao-header {
        background-color: #f0f2f6;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
        border-left: 4px solid #4CAF50;
    }
    
    .produto-row {
        background-color: #fafafa;
        padding: 8px;
        margin: 2px 0;
        border-radius: 3px;
        border-left: 2px solid #ddd;
    }
    
    .expander-content {
        margin-left: 20px;
    }

    /* Tabela de preço: mesmo tamanho do rótulo das métricas, texto completo */
    .loja-metrica-texto .loja-metrica-label,
    .loja-metrica-texto .loja-metrica-valor {
        font-size: 0.875rem;
        color: rgb(49, 51, 63);
        line-height: 1.35;
        font-weight: 400;
    }
    .loja-metrica-texto .loja-metrica-label {
        margin-bottom: 0.35rem;
    }
    .loja-metrica-texto .loja-metrica-valor {
        word-wrap: break-word;
        overflow-wrap: anywhere;
        white-space: normal;
    }
    .vo-promocao-meta {
        background: #f8f4fc;
        border: 1px solid #d4c4e8;
        border-radius: 6px;
        padding: 0.5rem 0.75rem;
        margin: 0.35rem 0 0.75rem 0;
        font-size: 0.875rem;
        color: rgb(49, 51, 63);
        text-align: left;
    }
    .vo-promocao-titulo {
        margin-top: 0.75rem;
        margin-bottom: 0.15rem;
        font-weight: 600;
        color: rgb(49, 51, 63);
        text-align: left;
    }
    .vo-promocao-legenda {
        text-align: left;
        font-size: 0.8125rem;
        color: rgb(112, 117, 128);
        margin: 0 0 0.5rem 0;
        font-style: italic;
    }
    .vo-promocao-resumo {
        text-align: left;
        font-size: 0.875rem;
        color: rgb(49, 51, 63);
        margin-bottom: 0.35rem;
    }
    .vo-promocao-tabela-wrap {
        overflow-x: auto;
        max-height: 420px;
        overflow-y: auto;
        border: 1px solid #e6e6e6;
        border-radius: 6px;
        margin-top: 0.25rem;
    }
    .vo-promocao-tabela {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.875rem;
        color: rgb(49, 51, 63);
    }
    .vo-promocao-tabela th,
    .vo-promocao-tabela td {
        text-align: left !important;
        padding: 0.45rem 0.75rem;
        border-bottom: 1px solid #ececec;
        vertical-align: top;
    }
    .vo-promocao-tabela th {
        font-weight: 600;
        background: #fafafa;
        position: sticky;
        top: 0;
        z-index: 1;
    }
    .vo-promocao-tabela tr:last-child td {
        border-bottom: none;
    }
    </style>
""", unsafe_allow_html=True)

# Configurações das marcas
MARCAS_CONFIG = {
    "Promoções Bendito": {
        "codfranqueador": 3082,
        "cor": "#FF6B6B",
        "nome_exibicao": "Momento Bendito",
    },
    "Promoções Espetto": {
        "codfranqueador": 3078,
        "cor": "#4ECDC4",
        "nome_exibicao": "Espetto Carioca",
    },
    "Promoções Mané": {
        "codfranqueador": 1428,
        "cor": "#95E1D3",
        "nome_exibicao": "Mané",
    },
    "Promoções Buteco Seu Rufino": {
        "codfranqueador": 3081,
        "cor": "#FFA500",
        "nome_exibicao": "Buteco Seu Rufino",
    },
}

CREDENCIAIS = {
    "usuario": "06266555794",
    "senha": "250913"
}


def _formatar_data_br(valor):
    """Formata data para exibição: dd/mm/aaaa."""
    if valor is None:
        return ""
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    return str(valor)

def _agora_brasilia_str(fmt="%d/%m/%Y %H:%M:%S"):
    """Data/hora atual em Brasília (America/Sao_Paulo) para exibição."""
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/Sao_Paulo")).strftime(fmt)
    except Exception:
        from datetime import timezone
        return datetime.now(timezone(timedelta(hours=-3))).strftime(fmt)



def _eh_promocoes_rede(nome_grupo):
    return "PROMOCOES REDE" in _normalizar_grupo(nome_grupo)


def listar_nomes_promocao_rede(df_marca):
    """Lista nomePromocao distintos onde o grupo é PROMOÇÕES REDE."""
    if df_marca.empty or "nomeGrupo" not in df_marca.columns:
        return []
    sub = df_marca[df_marca["nomeGrupo"].apply(_eh_promocoes_rede)]
    if sub.empty or "nomePromocao" not in sub.columns:
        return []
    return sorted(sub["nomePromocao"].dropna().astype(str).unique().tolist())


def codigos_produtos_promocao_rede(df_marca, nome_promocao):
    """Códigos de produto agregados da ação: PROMOÇÕES REDE + nomePromocao."""
    if df_marca.empty:
        return set()
    mask = (
        df_marca["nomeGrupo"].apply(_eh_promocoes_rede)
        & (df_marca["nomePromocao"].astype(str) == str(nome_promocao))
    )
    if "codigoProduto" not in df_marca.columns:
        return set()
    cods = df_marca.loc[mask, "codigoProduto"].dropna().unique()
    out = set()
    for c in cods:
        try:
            out.add(int(float(c)))
        except (TypeError, ValueError):
            continue
    return out


_JANELA_CARDAPIO_CODIGO_VENDA = 120


def _obter_cardapio_detalhado_loja(session, token, codfranqueador, codigo_loja):
    """GET relacao-cardapio-produto — itens com codigoProduto e valorVenda."""
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_CARDAPIO_PRODUTOS}"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        resp = session.get(
            url,
            params={"CodigoFranqueador": int(codfranqueador), "CodigoLoja": int(codigo_loja)},
            headers=headers,
            timeout=25,
        )
        if resp.status_code != 200:
            return []
        return [
            item
            for item in _extrair_lista_vo(resp.json())
            if isinstance(item, dict) and item.get("codigoProduto") is not None
        ]
    except Exception:
        return []


_MAX_FAIXAS_SKUS_UNITARIOS_PROMOCAO = 8


def _clusters_vendaveis_homogeneos_apos(cod, por_cod, max_itens=15):
    """
    Agrupa SKUs vendáveis consecutivos após o código de referência por valorVenda.
    Ex.: após 2396 → [[2397, 2398, 2399], [2400], [2401]].
    """
    clusters: list[list[int]] = []
    atual: list[int] = []
    valor_ref = None
    for c in range(cod + 1, cod + max_itens + 1):
        item = por_cod.get(c)
        if not item:
            break
        try:
            valor = float(item.get("valorVenda") or 0)
        except (TypeError, ValueError):
            break
        if valor <= 0:
            break
        if valor_ref is None or abs(valor - valor_ref) <= 0.01:
            valor_ref = valor
            atual.append(c)
            continue
        if atual:
            clusters.append(atual)
        atual = [c]
        valor_ref = valor
    if atual:
        clusters.append(atual)
    return clusters


def _codigos_vendaveis_expandidos_apos_referencia(clusters):
    """
    De cada cluster homogêneo (≥2 SKUs) fica o maior código; após um cluster multi,
    inclui no máximo um cluster unitário seguinte (faixa de preço extra da mesma ação).
    """
    out: set[int] = set()
    prev_multi = False
    for cluster in clusters:
        if len(cluster) >= 2:
            out.add(cluster[-1])
            prev_multi = True
        elif len(cluster) == 1 and prev_multi:
            out.add(cluster[0])
            prev_multi = False
        else:
            prev_multi = False
    return out


def _codigos_vendaveis_de_clusters_referencia(clusters):
    """
    Resolve SKUs vendáveis a partir dos clusters após o código de referência.

    - Somente unitários (Espetto, Rufino): inclui cada faixa de preço da ação
      quando há poucas faixas consecutivas (ex.: 1657/1658/1659 após 1656).
    - Primeiro cluster multi (Mané): principal de cada cluster homogêneo + até
      uma faixa unitária logo após o primeiro bloco multi (ex.: 2399 e 2400).
    - Unitários antes do primeiro multi (Champions): não expande; vendas ficam
      no código de referência (ex.: 2363).
    """
    if not clusters:
        return set()

    tem_multi = any(len(c) >= 2 for c in clusters)

    if not tem_multi:
        if len(clusters) <= _MAX_FAIXAS_SKUS_UNITARIOS_PROMOCAO:
            return {c[0] for c in clusters if c}
        return set()

    if len(clusters[0]) < 2:
        return set()

    return _codigos_vendaveis_expandidos_apos_referencia(clusters)


def _expandir_codigos_cardapio_loja(cods_base, cardapio_itens, janela=_JANELA_CARDAPIO_CODIGO_VENDA):
    """
    Quando o código da promoção é referência (valorVenda zerado), resolve os SKUs
    vendáveis no cardápio da loja — vale para qualquer marca (Mané, Espetto,
    Bendito, Rufino). Padrões suportados: faixas só unitárias, cluster homogêneo
    com faixa extra, ou venda no próprio código de referência (Champions).
    """
    _ = janela
    out = set(cods_base or [])
    if not out or not cardapio_itens:
        return out

    por_cod = {}
    for item in cardapio_itens:
        cod = _int_codigo_produto(item.get("codigoProduto"))
        if cod is not None:
            por_cod[cod] = item

    for cod in list(out):
        item = por_cod.get(cod)
        if not item:
            continue
        try:
            valor_ref = float(item.get("valorVenda") or 0)
        except (TypeError, ValueError):
            valor_ref = 0.0
        if valor_ref > 0:
            continue

        clusters = _clusters_vendaveis_homogeneos_apos(cod, por_cod)
        expandidos = _codigos_vendaveis_de_clusters_referencia(clusters)
        if not expandidos:
            continue
        # Mantém o código principal (a retaguarda agrupa por ele e ele pode ser
        # vendido diretamente em algumas lojas) e adiciona os SKUs vendáveis.
        out.update(expandidos)
    return out


def _refinar_codigos_acao_por_vendas(cods_base, principais, vendas):
    """
    Ajusta o conjunto de códigos de uma ação para contar APENAS o produto principal
    e seus componentes próprios — espelhando o agrupamento "por produto" da retaguarda.

    Usa a estrutura de combo das vendas (numLanctoItemPrincipal):
    - Acrescenta componentes que entram a R$ 0,00 ligados EXCLUSIVAMENTE às linhas da
      ação (ex.: ESPETTO BRASIL 1662, não consecutivo no cardápio), ignorando
      modificadores genéricos (ponto da carne, talher, molho), que são filhos de
      muitos produtos.
    - Quando a ação tem estrutura de combo, remove produtos vizinhos NÃO relacionados
      que a varredura de cardápio possa ter capturado (ex.: um BALDE listado logo
      abaixo da Copa), preservando sempre o(s) código(s) principal(is) selecionado(s).
    - Ações sem combo (ex.: Copa do Mané, que vende direto nos próprios códigos) não
      sofrem remoção: o conjunto base é mantido.
    """
    base = set(cods_base or [])
    principais = set(principais or [])
    if not base:
        return base

    child_owners = defaultdict(set)
    sold_paid = set()
    for v in vendas or []:
        if not _venda_nao_cancelada(v):
            continue
        lancto_code = {}
        for it in v.get("itens") or []:
            nl = it.get("numLancto")
            c = _int_codigo_produto(it.get("codProduto"))
            if nl is not None and c is not None:
                lancto_code[nl] = c
        for it in v.get("itens") or []:
            if not _item_conta_para_clique(it):
                continue
            c = _int_codigo_produto(it.get("codProduto"))
            if c is None:
                continue
            try:
                val = float(it.get("valUnitario") or 0)
            except (TypeError, ValueError):
                val = 0.0
            pai = it.get("numLanctoItemPrincipal")
            if pai in (0, None):
                if val > 0.01:
                    sold_paid.add(c)
            elif abs(val) < 0.01:
                child_owners[c].add(lancto_code.get(pai))

    exclusivos = {
        c: owners
        for c, owners in child_owners.items()
        if owners
        and all(o in base for o in owners)
        and c not in sold_paid
        and c not in base
    }

    combo_owners = set()
    for owners in exclusivos.values():
        combo_owners |= owners

    resultado = set(base) | set(exclusivos.keys())

    if combo_owners:
        for c in list(base):
            if (
                c not in principais
                and c in sold_paid
                and c not in combo_owners
                and c not in exclusivos
            ):
                resultado.discard(c)

    return resultado


def _mapa_codigos_cliques_por_loja(codfranqueador, cods_base, lojas_df, session, token):
    """codigoLoja -> conjunto de códigos (promoção ou código principal vendável por loja)."""
    mapa = {}
    for _, r in lojas_df.iterrows():
        try:
            cod_loja = int(r["codigoLoja"])
        except (TypeError, ValueError):
            continue
        cardapio = _obter_cardapio_detalhado_loja(session, token, codfranqueador, cod_loja)
        mapa[cod_loja] = _expandir_codigos_cardapio_loja(cods_base, cardapio)
    return mapa


PREFIX_OPCAO_VO_PROMOCAO = "[Categoria PROMOÇÃO] "
MIN_LOJAS_VO_AGREGADO = 3
FRACAO_LOJAS_VO_AGREGADO = 0.5


def _int_codigo_produto(valor):
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _codigos_de_linhas_vo(linhas):
    out = set()
    for ln in linhas or []:
        if not isinstance(ln, dict):
            continue
        cod = _int_codigo_produto(ln.get("Produto"))
        if cod is not None:
            out.add(cod)
    return out


def _codigos_promocoes_rede_union(df_marca):
    """Todos os códigos de produto em PROMOÇÕES REDE da marca (deduplicação VO)."""
    if df_marca.empty or "nomeGrupo" not in df_marca.columns:
        return set()
    sub = df_marca[df_marca["nomeGrupo"].apply(_eh_promocoes_rede)]
    if sub.empty or "codigoProduto" not in sub.columns:
        return set()
    out = set()
    for c in sub["codigoProduto"].dropna().unique():
        cod = _int_codigo_produto(c)
        if cod is not None:
            out.add(cod)
    return out


def _iter_unidades_vo_com_produtos(mapa_vo):
    for cod_loja, vo in sorted((mapa_vo or {}).items(), key=lambda x: int(x[0])):
        linhas = vo.get("linhas_retaguarda") or []
        if not linhas:
            continue
        nome_loja = vo.get("nome_loja") or "N/A"
        yield int(cod_loja), nome_loja, linhas


def _limiar_lojas_vo_agregado(total_lojas_com_vo):
    if total_lojas_com_vo <= 0:
        return MIN_LOJAS_VO_AGREGADO
    por_fracao = int(total_lojas_com_vo * FRACAO_LOJAS_VO_AGREGADO + 0.999)
    return max(MIN_LOJAS_VO_AGREGADO, por_fracao)


def _mapa_opcoes_vo_multiloja(mapa_vo, codigos_rede):
    """
    Agrupa produtos VO pelo rótulo Descrição Monitor (ou descrição do produto)
    quando o mesmo item aparece em várias lojas — promo com preço fechado na rede.
    """
    grupos = defaultdict(lambda: {"rotulo": "", "codigos": set(), "lojas": set()})
    lojas_com_vo = set()

    for cod_loja, _nome_loja, linhas in _iter_unidades_vo_com_produtos(mapa_vo):
        lojas_com_vo.add(cod_loja)
        for ln in linhas:
            if not isinstance(ln, dict):
                continue
            cod = _int_codigo_produto(ln.get("Produto"))
            if cod is None:
                continue
            monitor = (
                (ln.get("Descrição Monitor") or ln.get("Descrição do produto") or "")
                .strip()
            )
            if not monitor:
                monitor = f"Produto {cod}"
            chave = _normalizar_grupo(monitor)
            if not grupos[chave]["rotulo"]:
                grupos[chave]["rotulo"] = monitor
            grupos[chave]["codigos"].add(cod)
            grupos[chave]["lojas"].add(cod_loja)

    limiar = _limiar_lojas_vo_agregado(len(lojas_com_vo))
    opcoes = {}
    for info in grupos.values():
        if len(info["lojas"]) < limiar:
            continue
        cods = info["codigos"]
        if not cods:
            continue
        if cods.issubset(codigos_rede or set()):
            continue
        rotulo = info["rotulo"]
        label = f"{PREFIX_OPCAO_VO_PROMOCAO}{rotulo}"
        opcoes[label] = set(cods)
    return opcoes


def _mapa_opcoes_vo_por_produto(mapa_vo):
    """Um item por produto da categoria PROMOÇÃO (todas as lojas), chave = nome exibido."""
    grupos = defaultdict(lambda: {"rotulo": "", "codigos": set()})

    for _cod_loja, _nome_loja, linhas in _iter_unidades_vo_com_produtos(mapa_vo):
        for ln in linhas:
            if not isinstance(ln, dict):
                continue
            cod = _int_codigo_produto(ln.get("Produto"))
            if cod is None:
                continue
            rotulo = (
                (ln.get("Descrição do produto") or ln.get("Descrição Monitor") or "")
                .strip()
            )
            if not rotulo:
                rotulo = f"Produto {cod}"
            chave = _normalizar_grupo(rotulo)
            if not grupos[chave]["rotulo"]:
                grupos[chave]["rotulo"] = rotulo
            grupos[chave]["codigos"].add(cod)

    opcoes = {}
    for info in sorted(grupos.values(), key=lambda x: x["rotulo"].upper()):
        rotulo = info["rotulo"]
        cods = info["codigos"]
        if not cods:
            continue
        label = rotulo
        if label in opcoes and opcoes[label] != cods:
            for cod in sorted(cods):
                alt = f"{rotulo} ({cod})"
                opcoes[alt] = {cod}
        else:
            opcoes[label] = set(cods)
    return opcoes


def listar_opcoes_cliques_vo(mapa_vo):
    """Lista nomes de produtos VO (categoria PROMOÇÃO) para o bloco 2 de cliques."""
    return sorted(_mapa_opcoes_vo_por_produto(mapa_vo).keys())


def resolver_codigos_cliques_vo(mapa_vo, nome_produto):
    """Códigos Degust do produto VO selecionado (pode abranger várias lojas)."""
    if not nome_produto:
        return set()
    return set(_mapa_opcoes_vo_por_produto(mapa_vo).get(nome_produto, set()))


def listar_opcoes_cliques_promocao(df_marca, mapa_vo=None):
    """Bloco 1: apenas ações do grupo PROMOÇÕES DE REDE."""
    _ = mapa_vo
    return list(listar_nomes_promocao_rede(df_marca))


def resolver_codigos_cliques(df_marca, mapa_vo, nome_opcao):
    """Resolve codigoProduto para ação PROMOÇÕES DE REDE (bloco 1)."""
    _ = mapa_vo
    if not nome_opcao:
        return set()
    return codigos_produtos_promocao_rede(df_marca, nome_opcao)


def origem_opcao_cliques(nome_opcao):
    if not nome_opcao:
        return ""
    return "Promoções de rede"


def gerar_blocos_30_dias(data_inicio, data_fim):
    """
    Parte o intervalo em blocos de até 30 dias (API relatorio-vendas).
    Cada bloco: [início, fim] inclusive, no máximo 30 dias corridos.
    """
    if data_fim < data_inicio:
        return []
    blocos = []
    cur = data_inicio
    while cur <= data_fim:
        fim_bloco = min(cur + timedelta(days=29), data_fim)
        blocos.append((cur, fim_bloco))
        cur = fim_bloco + timedelta(days=1)
    return blocos


def _venda_nao_cancelada(venda):
    c = venda.get("cancelada")
    if c is None or c == "":
        return True
    return str(c).upper() not in ("S", "SIM", "1", "Y", "YES", "TRUE")


def _item_conta_para_clique(item):
    c = item.get("cancelado")
    if c is None or c == "":
        return True
    return str(c).upper() not in ("S", "SIM", "1", "Y", "YES", "TRUE")


def somar_cliques_em_vendas(vendas, produtos_set):
    """Soma quantidade dos itens cujo codProduto está em produtos_set (clique = venda)."""
    total = 0.0
    for v in vendas or []:
        if not _venda_nao_cancelada(v):
            continue
        for it in v.get("itens") or []:
            if not _item_conta_para_clique(it):
                continue
            cod = it.get("codProduto")
            if cod is None:
                continue
            try:
                c = int(cod)
            except (TypeError, ValueError):
                continue
            if c not in produtos_set:
                continue
            q = it.get("quantidade")
            try:
                total += float(q or 0)
            except (TypeError, ValueError):
                pass
    return total


def _normalizar_nom_usuario_venda(nome):
    if nome is None:
        return "(não informado)"
    texto = str(nome).strip()
    return texto if texto else "(não informado)"


def _chave_item_venda_clique(item):
    """Chave estável para cruzar relatório de vendas com o período sincronizado."""
    cod = item.get("codProduto")
    try:
        cod_int = int(cod)
    except (TypeError, ValueError):
        cod_int = 0
    momento = str(item.get("datHoraLancamento") or "")[:19]
    try:
        qtd = float(item.get("quantidade") or 0)
    except (TypeError, ValueError):
        qtd = 0.0
    return cod_int, momento, qtd


def _mapa_garcom_por_item_sync(vendas_sync):
    """(codProduto, datHoraLancamento, quantidade) -> codigoGarcom (API sincronizada)."""
    mapa = {}
    for venda in vendas_sync or []:
        if not _venda_nao_cancelada(venda):
            continue
        for item in venda.get("itens") or []:
            if not _item_conta_para_clique(item):
                continue
            garcom = item.get("codigoGarcom")
            try:
                garcom_int = int(garcom)
            except (TypeError, ValueError):
                continue
            if garcom_int <= 0:
                continue
            mapa[_chave_item_venda_clique(item)] = garcom_int
    return mapa


def _mapa_nome_por_garcom(vendas_rel, mapa_garcom_item):
    """
    Monta codigoGarcom -> nomUsuarioVenda cruzando relatório de vendas com a API
    sincronizada (mesmo item: produto + data/hora + quantidade).
    """
    nomes = {}
    for venda in vendas_rel or []:
        if not _venda_nao_cancelada(venda):
            continue
        for item in venda.get("itens") or []:
            if not _item_conta_para_clique(item):
                continue
            nome = str(item.get("nomUsuarioVenda") or "").strip()
            if not nome:
                continue
            garcom = mapa_garcom_item.get(_chave_item_venda_clique(item))
            if garcom:
                nomes[garcom] = nome
    return nomes


def _resolver_nome_garcom_item(item, mapa_garcom_item=None, mapa_nome_garcom=None):
    nome = str(item.get("nomUsuarioVenda") or "").strip()
    if nome:
        return nome
    if not mapa_garcom_item:
        return ""
    garcom = mapa_garcom_item.get(_chave_item_venda_clique(item))
    if not garcom:
        return ""
    if mapa_nome_garcom:
        nome_mapeado = mapa_nome_garcom.get(garcom)
        if nome_mapeado:
            return nome_mapeado
    return f"Garçom código {garcom}"


def somar_cliques_por_nom_usuario_venda(
    vendas,
    produtos_set,
    mapa_garcom_item=None,
    mapa_nome_garcom=None,
):
    """Soma quantidade por nomUsuarioVenda (itens da promoção, venda/item não cancelados)."""
    por_usuario = defaultdict(float)
    for v in vendas or []:
        if not _venda_nao_cancelada(v):
            continue
        for it in v.get("itens") or []:
            if not _item_conta_para_clique(it):
                continue
            cod = it.get("codProduto")
            if cod is None:
                continue
            try:
                c = int(cod)
            except (TypeError, ValueError):
                continue
            if c not in produtos_set:
                continue
            usuario = _resolver_nome_garcom_item(it, mapa_garcom_item, mapa_nome_garcom)
            usuario = _normalizar_nom_usuario_venda(usuario)
            q = it.get("quantidade")
            try:
                por_usuario[usuario] += float(q or 0)
            except (TypeError, ValueError):
                pass
    return dict(por_usuario)


def consultar_relatorio_vendas_list(session, token, cod_franqueador, cod_loja, d_ini, d_fim):
    """POST /api/venda/relatorio-vendas — retorna lista de vendas ou None."""
    url = f"{DEGUST_API_BASE}/api/venda/relatorio-vendas"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    body = {
        "codFranqueador": int(cod_franqueador),
        "codLoja": int(cod_loja),
        "dataInicial": d_ini.isoformat(),
        "dataFinal": d_fim.isoformat(),
        "tipoData": "C",
        "exibirVendasCanceladas": False,
    }
    try:
        r = session.post(url, json=body, headers=headers, timeout=120)
        if r.status_code != 200:
            return None
        vendas = r.json()
        return vendas if isinstance(vendas, list) else None
    except Exception:
        return None


def consultar_relatorio_vendas_agregados(
    session,
    token,
    cod_franqueador,
    cod_loja,
    d_ini,
    d_fim,
    produtos_set,
    mapa_garcom_item=None,
    mapa_nome_garcom=None,
):
    """POST /api/venda/relatorio-vendas — retorna (total cliques, cliques por nomUsuarioVenda) ou None."""
    url = f"{DEGUST_API_BASE}/api/venda/relatorio-vendas"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    body = {
        "codFranqueador": int(cod_franqueador),
        "codLoja": int(cod_loja),
        "dataInicial": d_ini.isoformat(),
        "dataFinal": d_fim.isoformat(),
        "tipoData": "C",
        "exibirVendasCanceladas": False,
    }
    try:
        r = session.post(url, json=body, headers=headers, timeout=120)
        if r.status_code != 200:
            return None
        vendas = r.json()
        if not isinstance(vendas, list):
            return None
        return (
            somar_cliques_em_vendas(vendas, produtos_set),
            somar_cliques_por_nom_usuario_venda(
                vendas,
                produtos_set,
                mapa_garcom_item=mapa_garcom_item,
                mapa_nome_garcom=mapa_nome_garcom,
            ),
        )
    except Exception:
        return None


def consultar_relatorio_vendas_sum(session, token, cod_franqueador, cod_loja, d_ini, d_fim, produtos_set):
    """Compat: retorna só a soma de cliques."""
    agg = consultar_relatorio_vendas_agregados(
        session, token, cod_franqueador, cod_loja, d_ini, d_fim, produtos_set
    )
    if agg is None:
        return None
    return agg[0]


def somar_cliques_por_garcom(vendas, produtos_set):
    """Soma quantidade por codigoGarcom (itens da promoção, venda/item não cancelados)."""
    por_garcom = defaultdict(float)
    for v in vendas or []:
        if not _venda_nao_cancelada(v):
            continue
        for it in v.get("itens") or []:
            if not _item_conta_para_clique(it):
                continue
            cod = it.get("codProduto")
            if cod is None:
                continue
            try:
                c = int(cod)
            except (TypeError, ValueError):
                continue
            if c not in produtos_set:
                continue
            g = it.get("codigoGarcom")
            if g is None:
                continue
            try:
                g_int = int(g)
            except (TypeError, ValueError):
                continue
            if g_int <= 0:
                continue
            q = it.get("quantidade")
            try:
                por_garcom[g_int] += float(q or 0)
            except (TypeError, ValueError):
                pass
    return dict(por_garcom)


def consultar_vendas_periodo_sincronizado(session, token, cod_franqueador, cod_loja, d_ini, d_fim):
    """POST /api/venda/relatorio-vendas-periodo-sincronizado — lista de vendas da loja."""
    url = f"{DEGUST_API_BASE}/api/venda/relatorio-vendas-periodo-sincronizado"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    body = {
        "codFranqueador": int(cod_franqueador),
        "dataInicial": f"{d_ini.isoformat()} 00:00:01",
        "dataFinal": f"{d_fim.isoformat()} 23:59:59",
        "tipo": "Lista",
        "listaDeLojas": str(int(cod_loja)),
        "tipoData": "V",
    }
    try:
        r = session.post(url, json=body, headers=headers, timeout=180)
        if r.status_code != 200:
            return None
        payload = r.json()
        if not isinstance(payload, dict):
            return None
        vendas = payload.get("vendas") or []
        return vendas if isinstance(vendas, list) else None
    except Exception:
        return None


def consultar_cliques_por_garcom_sinc(session, token, cod_franqueador, cod_loja, d_ini, d_fim, produtos_set):
    """POST /api/venda/relatorio-vendas-periodo-sincronizado — cliques por codigoGarcom."""
    vendas = consultar_vendas_periodo_sincronizado(
        session, token, cod_franqueador, cod_loja, d_ini, d_fim
    )
    if vendas is None:
        return None, None
    return somar_cliques_por_garcom(vendas, produtos_set), vendas


def _mesclar_contagem_garcom(acumulado, parcial):
    """Soma incremental de cliques por garçom (codigoGarcom) ao longo dos blocos de datas."""
    if not parcial:
        return acumulado
    for g, q in parcial.items():
        acumulado[g] += q
    return acumulado


def _mesclar_contagem_usuario_venda(acumulado, parcial):
    """Soma incremental de cliques por nomUsuarioVenda ao longo dos blocos de datas."""
    if not parcial:
        return acumulado
    for u, q in parcial.items():
        acumulado[u] += q
    return acumulado


def _formatar_ranking_engajados(por_contagem, rotulo_chave, top_n=None):
    """
    Ranking por cliques no período; uma linha por posição.
    top_n=None lista todos (do maior para o menor).
    """
    if not por_contagem:
        return "N/D"
    ranking = sorted(
        ((k, q) for k, q in por_contagem.items() if q > 0),
        key=lambda x: (-x[1], str(x[0])),
    )
    if top_n is not None:
        ranking = ranking[:top_n]
    if not ranking:
        return "N/D"
    ordinais = ("1º", "2º", "3º", "4º", "5º")
    linhas = []
    for i, (chave, qtd) in enumerate(ranking):
        qtd_fmt = int(qtd) if qtd == int(qtd) else round(qtd, 2)
        prefixo = ordinais[i] if i < len(ordinais) else f"{i + 1}º"
        rotulo_qtd = "Clique" if qtd_fmt == 1 else "Cliques"
        linhas.append(f"{prefixo}: {rotulo_chave(chave)} = {qtd_fmt} {rotulo_qtd}")
    return "\n".join(linhas)


def _formatar_top_garcons_engajados(por_garcom, top_n=None):
    """Ranking por codigoGarcom (PDV). Padrão: todos do período."""
    return _formatar_ranking_engajados(
        por_garcom,
        lambda cod: f"Código {cod}",
        top_n=top_n,
    )


def _formatar_ranking_nome_garcom(por_usuario):
    """Ranking por nomUsuarioVenda (Nome do garçom). Lista todos do período."""
    return _formatar_ranking_engajados(por_usuario, lambda nome: str(nome), top_n=None)


_ROTULO_TOTAL_CLIQUES_REDE = "TOTAL DE CLIQUES DAS LOJAS"
COL_CLIQUES_MARCA = "Marca"
COL_CLIQUES_NOME_LOJA = "Nome da loja"
COL_CLIQUES_ACUMULADO = "Acumulado (cliques)"
COL_CLIQUES_COD_GARCOM = "Código do garçom + engajado"
COL_CLIQUES_NOME_GARCOM = "Nome do Garçom"
COL_CLIQUES_TABELA_PRECO = "Tabela de preço atual da loja"
COL_CLIQUES_PERIODO = "Periodo"


def _nome_marca_exibicao(marca_interna):
    """Nome comercial da marca (ex.: Promoções Espetto → Espetto Carioca)."""
    cfg = MARCAS_CONFIG.get(str(marca_interna or ""), {})
    return cfg.get("nome_exibicao") or str(marca_interna or "N/A")


def _rotulos_periodo_exibicao(col_labels):
    """Mantém os rótulos originais de data (ex.: '13/05/2026 a 24/05/2026')."""
    return {lbl: lbl for lbl in col_labels}


def _renomear_colunas_exibicao_tabela_cliques(df, col_labels):
    mapa = {
        "marca": COL_CLIQUES_MARCA,
        "nomeLoja": COL_CLIQUES_NOME_LOJA,
        "Acumulado (cliques)": COL_CLIQUES_ACUMULADO,
        "Código do garçom + Engajado": COL_CLIQUES_COD_GARCOM,
        "Nome do garçom": COL_CLIQUES_NOME_GARCOM,
        "Tabela de Preço": COL_CLIQUES_TABELA_PRECO,
        **_rotulos_periodo_exibicao(col_labels),
    }
    return df.rename(columns={k: v for k, v in mapa.items() if k in df.columns})


def _estilizar_cabecalhos_tabela_cliques(df):
    return df.style.set_table_styles(
        [{"selector": "th", "props": [("font-weight", "bold")]}],
        overwrite=False,
    )


def _colunas_tabela_produtos_promocao(df):
    """Colunas visíveis da tabela de produtos (valorMix oculto)."""
    colunas_ordenadas = [
        "codigoProduto",
        "descricaoProduto",
        "domingo",
        "segunda",
        "terca",
        "quarta",
        "quinta",
        "sexta",
        "sabado",
        "restricaoHorario",
        "valorPromocionalMix",
    ]
    return [col for col in colunas_ordenadas if col in df.columns]


def _colunas_exibicao_tabela_cliques(df):
    """Colunas visíveis na tabela de cliques (oculta codigoLoja e coluna imediatamente antes do acumulado)."""
    cols = [c for c in df.columns if c != "codigoLoja"]
    if COL_CLIQUES_ACUMULADO in cols:
        idx = cols.index(COL_CLIQUES_ACUMULADO)
        if idx > 0:
            cols = cols[: idx - 1] + cols[idx:]
    return cols


def montar_tabela_cliques_promocao_rede(
    codfranqueador,
    df_marca,
    nome_promocao,
    data_inicio,
    data_fim,
    max_workers=6,
    progress_bar=None,
    status_label=None,
    mapa_vo=None,
    produtos_set=None,
):
    """
    Por loja: colunas = um bloco de até 30 dias cada + Acumulado (cliques).
    Retorna (DataFrame, mensagem_erro). mensagem_erro só se falha grosseira.
    """
    if produtos_set is None:
        produtos_set = resolver_codigos_cliques(df_marca, mapa_vo, nome_promocao)
    if not produtos_set:
        return None, (
            "Nenhum produto encontrado para esta opção "
            "(Promoções de rede ou categoria PROMOÇÃO na venda orientada)."
        )

    if "codigoLoja" not in df_marca.columns or "nomeLoja" not in df_marca.columns:
        return None, "DataFrame sem codigoLoja ou nomeLoja."

    lojas_df = df_marca[["codigoLoja", "nomeLoja"]].drop_duplicates().sort_values("codigoLoja")
    blocos = gerar_blocos_30_dias(data_inicio, data_fim)
    if not blocos:
        return None, "Intervalo de datas inválido."

    if "marca" in df_marca.columns and not df_marca["marca"].dropna().empty:
        nome_marca = _nome_marca_exibicao(df_marca["marca"].dropna().iloc[0])
    else:
        nome_marca = "N/A"

    col_labels = [
        f"{_formatar_data_br(a)} a {_formatar_data_br(b)}" for a, b in blocos
    ]

    rows_by_cod = {}
    garcom_por_loja = defaultdict(lambda: defaultdict(float))
    usuario_por_loja = defaultdict(lambda: defaultdict(float))
    for _, r in lojas_df.iterrows():
        cod_loja_row = int(r["codigoLoja"])
        rows_by_cod[cod_loja_row] = {
            "codigoLoja": cod_loja_row,
            "marca": nome_marca,
            "nomeLoja": r.get("nomeLoja", "N/A"),
        }
        for lbl in col_labels:
            rows_by_cod[cod_loja_row][lbl] = 0.0

    total_ops = len(blocos) * len(lojas_df)
    done = 0
    thread_local = threading.local()

    def _sess():
        if not hasattr(thread_local, "session"):
            thread_local.session = requests.Session()
        return thread_local.session

    with requests.Session() as main_session:
        token = autenticar(codfranqueador, session=main_session)
        if not token:
            return None, "Falha na autenticação."

        mapa_cods_loja = _mapa_codigos_cliques_por_loja(
            codfranqueador, produtos_set, lojas_df, main_session, token
        )

        for bidx, (di, df_end) in enumerate(blocos):
            lbl = col_labels[bidx]

            def _work(cod_loja):
                loja_cod = int(cod_loja)
                cods_loja = mapa_cods_loja.get(loja_cod) or produtos_set
                sess = _sess()
                vendas_sync = consultar_vendas_periodo_sincronizado(
                    sess, token, codfranqueador, loja_cod, di, df_end
                ) or []
                vendas_rel = consultar_relatorio_vendas_list(
                    sess, token, codfranqueador, loja_cod, di, df_end
                )
                if vendas_rel is None:
                    soma, por_usuario = 0.0, {}
                else:
                    cods_loja = _refinar_codigos_acao_por_vendas(
                        cods_loja, produtos_set, vendas_rel
                    )
                    mapa_garcom_item = _mapa_garcom_por_item_sync(vendas_sync)
                    mapa_nome_garcom = _mapa_nome_por_garcom(vendas_rel, mapa_garcom_item)
                    soma = somar_cliques_em_vendas(vendas_rel, cods_loja)
                    por_usuario = somar_cliques_por_nom_usuario_venda(
                        vendas_rel,
                        cods_loja,
                        mapa_garcom_item=mapa_garcom_item,
                        mapa_nome_garcom=mapa_nome_garcom,
                    )
                por_garcom = somar_cliques_por_garcom(vendas_sync, cods_loja)
                return loja_cod, soma, por_usuario, por_garcom or {}

            workers = max(1, min(max_workers, len(lojas_df)))
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futs = [
                    executor.submit(_work, r["codigoLoja"]) for _, r in lojas_df.iterrows()
                ]
                for fut in as_completed(futs):
                    loja_cod, soma, por_usuario, por_garcom = fut.result()
                    rows_by_cod[loja_cod][lbl] = soma
                    _mesclar_contagem_garcom(garcom_por_loja[loja_cod], por_garcom)
                    _mesclar_contagem_usuario_venda(usuario_por_loja[loja_cod], por_usuario)
                    done += 1
                    if progress_bar is not None:
                        progress_bar.progress(min(done / max(total_ops, 1), 1.0))
                    if status_label is not None:
                        status_label.text(
                            f"Relatório de vendas + garçom: bloco {bidx + 1}/{len(blocos)} "
                            f"({lbl}) — {done}/{total_ops}"
                        )

        mapa_tabela_preco = _mapa_tabela_preco_por_loja(
            df_marca,
            list(rows_by_cod.keys()),
            cliente_http=main_session,
            token=token,
            codfranqueador=codfranqueador,
        )

    col_garcom = "Código do garçom + Engajado"
    col_nome_garcom = "Nome do garçom"
    col_tabela_preco = "Tabela de Preço"
    for c in rows_by_cod:
        acc = sum(rows_by_cod[c][lbl] for lbl in col_labels)
        rows_by_cod[c]["Acumulado (cliques)"] = acc
        rows_by_cod[c][col_garcom] = _formatar_top_garcons_engajados(garcom_por_loja.get(c))
        rows_by_cod[c][col_nome_garcom] = _formatar_ranking_nome_garcom(usuario_por_loja.get(c))
        rows_by_cod[c][col_tabela_preco] = mapa_tabela_preco.get(int(c), "N/A")

    ordem = (
        ["codigoLoja", "marca", "nomeLoja"]
        + col_labels
        + ["Acumulado (cliques)", col_nome_garcom, col_tabela_preco]
    )
    df_out = pd.DataFrame(list(rows_by_cod.values()))[ordem]
    df_out = df_out.sort_values(
        by="Acumulado (cliques)", ascending=False, ignore_index=True
    )

    # Linha de totais gerais (soma por coluna de período + acumulado).
    linha_total = {"codigoLoja": "", "marca": "", "nomeLoja": _ROTULO_TOTAL_CLIQUES_REDE}
    for lbl in col_labels:
        linha_total[lbl] = float(df_out[lbl].sum())
    linha_total["Acumulado (cliques)"] = float(df_out["Acumulado (cliques)"].sum())
    linha_total[col_nome_garcom] = ""
    linha_total[col_tabela_preco] = ""
    df_out = pd.concat([df_out, pd.DataFrame([linha_total])], ignore_index=True)

    # Garantir que colunas de contagem apareçam como inteiros (antes o app mostrava assim).
    for lbl in col_labels:
        df_out[lbl] = (
            pd.to_numeric(df_out[lbl], errors="coerce")
            .fillna(0)
            .round(0)
            .astype(int)
        )
    if "Acumulado (cliques)" in df_out.columns:
        df_out["Acumulado (cliques)"] = (
            pd.to_numeric(df_out["Acumulado (cliques)"], errors="coerce")
            .fillna(0)
            .round(0)
            .astype(int)
        )
    df_out = _renomear_colunas_exibicao_tabela_cliques(df_out, col_labels)

    return df_out, None

def _session_flag_true_callback(flag_key: str):
    """on_click do Streamlit: define flag antes do rerun (evita expander fechado no mobile)."""
    def _on_click():
        st.session_state[flag_key] = True
    return _on_click


def _render_bloco_cliques_por_loja(
    marca,
    df_marca,
    titulo_expander,
    texto_markdown,
    caption_help,
    prefixo_key,
    opcoes,
    label_selectbox,
    resolver_produtos_fn,
    origem_label,
    mensagem_vazio,
    excel_prefix,
):
    """Expander reutilizável: selectbox + consulta de cliques por loja e período."""
    _exp_key = f"ui_exp_cliques_{prefixo_key}_{marca}"
    _chave_df = f"cliques_{prefixo_key}_df_{marca}"
    _exp_aberto = (
        st.session_state.get(_exp_key, False)
        or (
            _chave_df in st.session_state
            and st.session_state.get(_chave_df) is not None
        )
    )
    with st.expander(titulo_expander, expanded=_exp_aberto):
        st.markdown(texto_markdown)
        st.caption(caption_help)
        if not opcoes:
            st.info(mensagem_vazio)
            return
        nome_sel = st.selectbox(
            label_selectbox,
            opcoes,
            key=f"sel_cliques_{prefixo_key}_{marca}",
        )
        c_a, c_b, c_c = st.columns(3)
        with c_a:
            d_ini_acao = st.date_input(
                "Início da ação",
                value=date.today() - timedelta(days=29),
                format="DD/MM/YYYY",
                key=f"dini_cliques_{prefixo_key}_{marca}",
            )
        with c_b:
            d_fim_analise = st.date_input(
                "Último dia da análise",
                value=date.today(),
                format="DD/MM/YYYY",
                key=f"dfim_cliques_{prefixo_key}_{marca}",
            )
        with c_c:
            max_wr = st.number_input(
                "Qtda de consultas por loja ao mesmo tempo (Limite 6)",
                min_value=1,
                max_value=6,
                value=2,
                key=f"max_wr_cliques_{prefixo_key}_{marca}",
            )

        n_prods = len(resolver_produtos_fn(nome_sel))
        st.caption(
            f"Origem: **{origem_label}** · Produtos agregados: **{n_prods}** código(s)."
        )

        if st.button(
            "Consultar cliques por loja e período",
            key=f"btn_cliques_{prefixo_key}_{marca}",
            on_click=_session_flag_true_callback(_exp_key),
            use_container_width=True,
        ):
            if d_fim_analise < d_ini_acao:
                st.error("A data final deve ser maior ou igual à data de início da ação.")
            else:
                codfranqueador = MARCAS_CONFIG[marca]["codfranqueador"]
                produtos_set = resolver_produtos_fn(nome_sel)
                prog = st.progress(0)
                status_txt = st.empty()
                with st.spinner(
                    "Consultando relatório de vendas e garçom por loja (pode levar alguns minutos)…"
                ):
                    df_cliques, err = montar_tabela_cliques_promocao_rede(
                        codfranqueador,
                        df_marca,
                        nome_sel,
                        d_ini_acao,
                        d_fim_analise,
                        max_workers=int(max_wr),
                        progress_bar=prog,
                        status_label=status_txt,
                        produtos_set=produtos_set,
                    )
                prog.empty()
                status_txt.empty()
                if err:
                    st.error(err)
                elif df_cliques is not None and not df_cliques.empty:
                    st.session_state[_chave_df] = df_cliques
                    st.session_state[f"cliques_{prefixo_key}_meta_{marca}"] = {
                        "promocao": nome_sel,
                        "inicio": _formatar_data_br(d_ini_acao),
                        "fim": _formatar_data_br(d_fim_analise),
                        "gerado_em": _agora_brasilia_str(),
                    }
                    st.success("Consulta concluída.")
                else:
                    st.warning("Nenhum dado retornado.")

        if _chave_df in st.session_state and st.session_state[_chave_df] is not None:
            meta = st.session_state.get(f"cliques_{prefixo_key}_meta_{marca}", {})
            if meta:
                _ge = meta.get("gerado_em")
                _suf = f" · Consulta gerada em {_ge}" if _ge else ""
                st.caption(
                    f"Última consulta: **{meta.get('promocao', '')}** — "
                    f"{meta.get('inicio', '')} a {meta.get('fim', '')}{_suf}"
                )
            _df_cliq = st.session_state[_chave_df].copy()
            _rank = _df_cliq[
                _df_cliq[COL_CLIQUES_NOME_LOJA].astype(str)
                != _ROTULO_TOTAL_CLIQUES_REDE
            ]
            _top_loja = None
            if (
                not _rank.empty
                and COL_CLIQUES_ACUMULADO in _rank.columns
                and COL_CLIQUES_NOME_LOJA in _rank.columns
            ):
                _acc = pd.to_numeric(
                    _rank[COL_CLIQUES_ACUMULADO], errors="coerce"
                ).fillna(0)
                _top_loja = str(_rank.loc[_acc.idxmax(), COL_CLIQUES_NOME_LOJA])
            if _top_loja:
                st.markdown(
                    f"Loja com mais aderência na ação até o momento: **{_top_loja}** 🏆"
                )
            _cols_show = _colunas_exibicao_tabela_cliques(_df_cliq)
            st.dataframe(
                _estilizar_cabecalhos_tabela_cliques(_df_cliq[_cols_show]),
                use_container_width=True,
                hide_index=True,
            )
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                st.session_state[_chave_df].to_excel(
                    writer, index=False, sheet_name="Cliques"
                )
                ws = writer.sheets["Cliques"]
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            buf.seek(0)
            st.download_button(
                label="⬇️ Baixar tabela de cliques (Excel)",
                data=buf,
                file_name=f"{excel_prefix}_{marca.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_cliques_{prefixo_key}_{marca}",
            )


def _nome_promocoes_unidade_loja(nome_loja):
    return f"PROMOÇÕES - {str(nome_loja or '').upper().strip()}"


_URL_VO_CONFIG = "/api/venda-orientada/consultar-venda-orientada"
_URL_VO_PRODUTO = "/api/venda-orientada/consultar-produto-por-grupo-venda-orientada"
_URL_LOJA_DETALHE = "/api/loja/loja"
_URL_CARDAPIO_PRODUTOS = "/api/produto/relacao-cardapio-produto"
_GRUPO_VO_PROMOCAO = "PROMOCAO"
_VO_PROMOCAO_MIN_PRODUTOS = 5
_VO_SCAN_MAX = 400


def _extrair_lista_vo(dados):
    if dados is None:
        return []
    if isinstance(dados, list):
        return dados
    if isinstance(dados, dict):
        for chave in ("data", "produtos", "itens", "items", "resultado", "content"):
            val = dados.get(chave)
            if isinstance(val, list):
                return val
        return [dados]
    return []


def _grupo_vo_eh_promocao(item):
    desc = str(item.get("grupoDescricao") or item.get("descricaoGrupo") or "").strip().upper()
    return desc == _GRUPO_VO_PROMOCAO or desc == "PROMOÇÃO" or desc == "PROMOÇOES"


def _obter_nome_venda_orientada_loja(session, token, codfranqueador, codigo_loja):
    """Nome da VO no cadastro da loja (ex.: OUT_2025 EC RECREIO) — GET /api/loja/loja."""
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_LOJA_DETALHE}"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        resp = session.get(
            url,
            params={"CodigoFranqueador": int(codfranqueador), "CodigoLoja": int(codigo_loja)},
            headers=headers,
            timeout=15,
        )
        if resp.status_code != 200:
            return ""
        cfg = (resp.json() or {}).get("configuracaoVenda") or {}
        return str(cfg.get("configuracaoVendaOrientada") or "").strip()
    except Exception:
        return ""


def _obter_ids_cardapio_loja(session, token, codfranqueador, codigo_loja):
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_CARDAPIO_PRODUTOS}"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        resp = session.get(
            url,
            params={"CodigoFranqueador": int(codfranqueador), "CodigoLoja": int(codigo_loja)},
            headers=headers,
            timeout=25,
        )
        if resp.status_code != 200:
            return frozenset()
        return frozenset(
            item.get("codigoProduto")
            for item in _extrair_lista_vo(resp.json())
            if item.get("codigoProduto") is not None
        )
    except Exception:
        return frozenset()


def _carregar_promocao_por_codigo_vo(session, token, codfranqueador, max_vo=_VO_SCAN_MAX):
    """Mapa codigo vendaOrientada (int) -> ids de produtos do grupo PROMOCAO."""
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_VO_PRODUTO}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    mapa = {}

    def _probe(codigo_vo):
        try:
            resp = session.post(
                url,
                json={"codigoFranquia": int(codfranqueador), "vendaOrientada": int(codigo_vo)},
                headers=headers,
                timeout=12,
            )
            if resp.status_code != 200:
                return codigo_vo, frozenset()
            prom = frozenset(
                item.get("produto")
                for item in _extrair_lista_vo(resp.json())
                if _grupo_vo_eh_promocao(item) and item.get("produto") is not None
            )
            return codigo_vo, prom
        except Exception:
            return codigo_vo, frozenset()

    workers = min(8, max(1, max_vo))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        for codigo_vo, prom in executor.map(_probe, range(1, max_vo + 1)):
            if prom:
                mapa[int(codigo_vo)] = prom
    return mapa


def _config_casa_loja(config_nome, nome_loja):
    cfg = _normalizar_grupo(config_nome).replace(" ", "")
    loja = _normalizar_grupo(nome_loja).replace(" ", "")
    if not cfg or not loja:
        return False
    return loja in cfg or cfg.endswith(loja)


def _mapear_configuracao_vo_franquia(session, token, codfranqueador, lojas, mapa_prom_vo):
    """
    Associa configuracaoVendaOrientada (nome) -> codigo vendaOrientada (int).
    Usa GET /api/loja/loja + overlap cardapio x grupo PROMOCAO (1 VO por config).
    """
    config_lojas = {}
    for loja in lojas or []:
        try:
            codigo_loja = int(loja["codigoLoja"])
        except (TypeError, ValueError):
            continue
        if codigo_loja == 999:
            continue
        nome_loja = loja.get("nomeLoja") or ""
        config_nome = _obter_nome_venda_orientada_loja(
            session, token, codfranqueador, codigo_loja
        )
        if not config_nome:
            continue
        cardapio_ids = _obter_ids_cardapio_loja(session, token, codfranqueador, codigo_loja)
        if not cardapio_ids:
            continue
        config_lojas.setdefault(config_nome, []).append(
            {"codigo_loja": codigo_loja, "nome_loja": nome_loja, "cardapio": cardapio_ids}
        )

    if not config_lojas or not mapa_prom_vo:
        return {}

    edges = []
    for config_nome, entries in config_lojas.items():
        for codigo_vo, prom in mapa_prom_vo.items():
            if len(prom) < _VO_PROMOCAO_MIN_PRODUTOS:
                continue
            overlaps = []
            bonus_nome = 0.0
            for entry in entries:
                card = entry["cardapio"]
                overlaps.append(len(prom & card) / len(prom))
                if _config_casa_loja(config_nome, entry["nome_loja"]):
                    bonus_nome = 0.02
            if not overlaps:
                continue
            overlap = max(overlaps)
            if overlap >= 0.80:
                miss = min(len(prom - entry["cardapio"]) for entry in entries)
                score = overlap + bonus_nome + (0.06 * miss)
                edges.append((score, overlap, len(prom), codigo_vo, config_nome))

    edges.sort(reverse=True)
    mapa_config = {}
    used_vo = set()
    for score, overlap, n_prom, codigo_vo, config_nome in edges:
        if config_nome in mapa_config or codigo_vo in used_vo:
            continue
        mapa_config[config_nome] = int(codigo_vo)
        used_vo.add(int(codigo_vo))

    return mapa_config


def _consultar_venda_orientada_config(session, token, codfranqueador, codigo_vo):
    """consultar-venda-orientada: configuração da VO (complementar; pode retornar vazio)."""
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_VO_CONFIG}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = session.post(
            url,
            json={"codigoFranquia": int(codfranqueador), "codigoVendaOrientada": int(codigo_vo)},
            headers=headers,
            timeout=15,
        )
        if resp.status_code == 200:
            return _extrair_lista_vo(resp.json())
    except Exception:
        pass
    return []


def _consultar_produtos_grupo_vo(session, token, codfranqueador, codigo_vo):
    """consultar-produto-por-grupo-venda-orientada — filtra grupo PROMOCAO."""
    url = f"{DEGUST_API_BASE.rstrip('/')}{_URL_VO_PRODUTO}"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        resp = session.post(
            url,
            json={"codigoFranquia": int(codfranqueador), "vendaOrientada": int(codigo_vo)},
            headers=headers,
            timeout=15,
        )
        if resp.status_code != 200:
            return []
        return [
            item
            for item in _extrair_lista_vo(resp.json())
            if isinstance(item, dict) and _grupo_vo_eh_promocao(item)
        ]
    except Exception:
        return []


def _valor_campo_vo(item, *chaves, default=""):
    for chave in chaves:
        val = item.get(chave)
        if val is not None and str(val).strip() not in ("", "None", "N/A"):
            return str(val).strip()
    return default


def _linha_retaguarda_vo(item):
    produto = _valor_campo_vo(
        item, "produto", "codigoProduto", "idProduto", "codigo", "id", default=""
    )
    descricao = _valor_campo_vo(
        item,
        "produtoDescricao",
        "descricao do produto",
        "descricaoProduto",
        "nomeProduto",
        "descricao",
        default="",
    )
    grupo = _valor_campo_vo(
        item, "grupo", "nomeGrupoVendaOrientada", "grupoVendaOrientada", default=""
    )
    desc_grupo = _valor_campo_vo(
        item, "grupoDescricao", "descricaoGrupo", "descricao do grupo", default=_GRUPO_VO_PROMOCAO
    )
    monitor = _valor_campo_vo(
        item, "descricaoMonitor", "Descrição Monitor", default=descricao
    )
    linha = {
        "Produto": produto,
        "Descrição do produto": descricao,
        "Grupo": grupo,
        "Descrição do Grupo": desc_grupo,
        "Descrição Monitor": monitor,
    }
    for dia in ("domingo", "segunda", "terca", "quarta", "quinta", "sexta", "sabado"):
        linha[dia] = _valor_campo_vo(item, dia, default="")
    linha["restricaoHorario"] = _valor_campo_vo(item, "restricaoHorario", default="")
    linha["valorMix"] = _valor_campo_vo(item, "valorMix", default="")
    linha["valorPromocionalMix"] = _valor_campo_vo(item, "valorPromocionalMix", default="")
    return linha


def _processar_loja_vo_promocao(token, codfranqueador, loja, mapa_config_vo):
    thread_local = threading.local()

    def _session():
        if not hasattr(thread_local, "session"):
            thread_local.session = requests.Session()
        return thread_local.session

    session = _session()
    codigo_loja = int(loja["codigoLoja"])
    nome_loja = loja.get("nomeLoja") or "N/A"
    nome_vo = _obter_nome_venda_orientada_loja(session, token, codfranqueador, codigo_loja)
    if not nome_vo:
        return None

    codigo_vo = (mapa_config_vo or {}).get(nome_vo)
    if codigo_vo is None:
        return None

    produtos = _consultar_produtos_grupo_vo(session, token, codfranqueador, codigo_vo)
    _consultar_venda_orientada_config(session, token, codfranqueador, codigo_vo)
    linhas = [_linha_retaguarda_vo(p) for p in produtos if isinstance(p, dict)]
    if not linhas:
        return None

    return {
        "codigo_loja": codigo_loja,
        "nome_loja": nome_loja,
        "linhas_retaguarda": linhas,
        "venda_orientada": int(codigo_vo),
        "venda_orientada_rotulo": nome_vo,
        "metodo_venda_orientada": (
            f"consultar-venda-orientada + consultar-produto-por-grupo-venda-orientada "
            f"(config={nome_vo}, VO={codigo_vo})"
        ),
    }


def _eh_promocao_unidade_por_nome(nome_promocao, nome_loja):
    """Identifica a seção PROMOÇÕES - {LOJA} do dashboard."""
    if not nome_promocao or not nome_loja:
        return False
    alvo = _normalizar_grupo(_nome_promocoes_unidade_loja(nome_loja))
    atual = _normalizar_grupo(nome_promocao)
    return atual == alvo or atual.startswith("PROMOCOES -")


@st.cache_data(ttl=300)
def carregar_mapa_categoria_vo_por_loja(codfranqueador, total_lojas):
    """Produtos VO PROMOCAO (CATEGORIA PROMOÇÃO) indexados por codigoLoja — API PRD do dashboard."""
    _ = total_lojas
    try:
        with requests.Session() as session:
            token = autenticar(int(codfranqueador), session=session)
            if not token:
                return {}
            lojas = obter_lojas(token, int(codfranqueador), session=session)
            if not lojas:
                return {}

            mapa_prom_vo = _carregar_promocao_por_codigo_vo(session, token, int(codfranqueador))
            if not mapa_prom_vo:
                return {}

            mapa_config_vo = _mapear_configuracao_vo_franquia(
                session, token, int(codfranqueador), lojas, mapa_prom_vo
            )
            if not mapa_config_vo:
                return {}

            workers = min(8, max(1, len(lojas)))
            unidades = []

            with ThreadPoolExecutor(max_workers=workers) as executor:
                futuros = {
                    executor.submit(
                        _processar_loja_vo_promocao, token, int(codfranqueador), loja, mapa_config_vo
                    ): loja
                    for loja in lojas
                }
                for futuro in as_completed(futuros):
                    try:
                        resultado = futuro.result()
                        if resultado:
                            unidades.append(resultado)
                    except Exception:
                        continue
    except Exception:
        return {}

    mapa = {}
    for unidade in unidades:
        try:
            cod = int(unidade.get("codigo_loja"))
        except (TypeError, ValueError):
            continue
        mapa[cod] = unidade
    return mapa


def _garantir_secao_promocoes_loja_vo(grupos_lojas, mapa_categoria_vo):
    """Garante bloco PROMOÇÕES - LOJA quando há produtos da CATEGORIA PROMOÇÃO."""
    codigos = set((mapa_categoria_vo or {}).keys())
    for cod_loja in codigos:
        vo = (mapa_categoria_vo or {}).get(cod_loja) or {}
        linhas_cat = vo.get("linhas_retaguarda") or []
        if not linhas_cat:
            continue
        nome_loja = vo.get("nome_loja") or "N/A"
        chave_loja = f"{cod_loja} - {nome_loja}"
        if chave_loja not in grupos_lojas:
            grupos_lojas[chave_loja] = {
                "info_loja": {
                    "codigoLoja": cod_loja,
                    "nomeLoja": nome_loja,
                    "marca": "N/A",
                    "tabelaDePreco": "N/A",
                },
                "promocoes": {},
            }
        nome_promo = _nome_promocoes_unidade_loja(nome_loja)
        if nome_promo not in grupos_lojas[chave_loja]["promocoes"]:
            grupos_lojas[chave_loja]["promocoes"][nome_promo] = {
                "info_promocao": {
                    "nomePromocao": nome_promo,
                    "promocaoAtiva": "Sim",
                    "nomeGrupo": "PROMOÇÕES DA UNIDADE",
                    "sequencia": None,
                },
                "produtos": [],
                "categorias": {},
            }


def _html_tabela_vo_promocao(df):
    """Renderiza tabela VO com alinhamento à esquerda (Streamlit 1.31 ignora Styler no grid)."""
    headers = "".join(f"<th>{html.escape(str(col))}</th>" for col in df.columns)
    body_rows = []
    for _, row in df.iterrows():
        cells = "".join(
            f"<td>{html.escape('' if pd.isna(val) else str(val))}</td>"
            for val in row
        )
        body_rows.append(f"<tr>{cells}</tr>")
    tbody = "".join(body_rows)
    return (
        '<div class="vo-promocao-tabela-wrap">'
        f'<table class="vo-promocao-tabela"><thead><tr>{headers}</tr></thead>'
        f"<tbody>{tbody}</tbody></table></div>"
    )


def _exibir_tabela_promocoes_loja(nome_loja, produtos):
    """Tabela da seção PROMOÇÕES - LOJA com dados do endpoint consultar-promocoes."""
    titulo = f'📋 PROMOÇÕES - {str(nome_loja or "").upper()}'
    st.markdown(f"**{titulo}**")
    if not produtos:
        st.info(f"Nenhum produto encontrado em PROMOÇÕES - {str(nome_loja or '').upper()}.")
        return

    df_produtos = pd.DataFrame(produtos)
    colunas_existentes = _colunas_tabela_produtos_promocao(df_produtos)
    df_produtos_ordenado = df_produtos[colunas_existentes]
    st.dataframe(
        df_produtos_ordenado,
        use_container_width=True,
        height=min(400, len(df_produtos) * 35 + 50),
    )


def _exibir_categoria_promocao_vo(nome_loja, vo_dados=None):
    """CATEGORIA PROMOÇÃO: produtos VO criados (API PRD), distintos do cardápio com desconto."""
    st.markdown(
        '<div class="vo-promocao-titulo">📋 CATEGORIA PROMOÇÃO</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="vo-promocao-legenda">Aqui mostra os produtos de promoção que foram criados. '
        "Que não são produtos do cardápio que sofreram desconto</div>",
        unsafe_allow_html=True,
    )

    linhas = (vo_dados or {}).get("linhas_retaguarda") or []
    if not linhas:
        st.info("Nenhum produto encontrado na categoria PROMOÇÃO (venda orientada — grupo PROMOCAO).")
        return

    n = len(linhas)
    vo = vo_dados.get("venda_orientada", "N/A")
    metodo = vo_dados.get("metodo_venda_orientada", "")
    rotulo = vo_dados.get("venda_orientada_rotulo", "")

    st.markdown(
        f'<div class="vo-promocao-resumo">🟢 Ativo · <b>{n} produtos</b> · '
        f"<b>Grupo:</b> PROMOCAO · <b>VO:</b> <code>{html.escape(str(vo))}</code></div>",
        unsafe_allow_html=True,
    )

    meta = (
        f"Venda Orientada: <b>{html.escape(str(rotulo))}</b> [{vo}]"
        f" · Descoberta: <b>{html.escape(str(metodo))}</b>"
    )
    st.markdown(f'<div class="vo-promocao-meta">{meta}</div>', unsafe_allow_html=True)

    df_vo = pd.DataFrame(linhas)
    cols = [c for c in COLUNAS_VO_PROMOCAO_EXIBICAO if c in df_vo.columns]
    st.markdown(_html_tabela_vo_promocao(df_vo[cols]), unsafe_allow_html=True)


def _normalizar_grupo(valor):
    """Normaliza nome de grupo para comparação consistente."""
    texto = str(valor or "").strip().upper()
    return (
        texto
        .replace("Ç", "C")
        .replace("Ã", "A")
        .replace("Á", "A")
        .replace("Â", "A")
        .replace("À", "A")
        .replace("É", "E")
        .replace("Ê", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ô", "O")
        .replace("Õ", "O")
        .replace("Ú", "U")
    )

def _grupo_deve_exibir_sequencia(nome_grupo):
    grupo = _normalizar_grupo(nome_grupo)
    return (
        "HAPPY HOUR" in grupo
        or "PROMOCOES DA UNIDADE" in grupo
        or "PROMOCOES REDE" in grupo
    )

def _extrair_sequencia_promocao(row, nome_grupo):
    """Retorna sequência quando o grupo exige exibição da ordem da promoção."""
    if not _grupo_deve_exibir_sequencia(nome_grupo):
        return None
    sequencia = row.get("sequencia", None)
    return sequencia if sequencia not in ("", "None") else None

def autenticar(codfranqueador, session=None):
    """Realiza autenticação na API do Degust"""
    url_auth = f"{DEGUST_API_BASE}/api/usuario/autenticar"
    
    credenciais = {
        "usuario": CREDENCIAIS["usuario"],
        "senha": CREDENCIAIS["senha"],
        "codigoFranqueador": codfranqueador
    }
    cliente_http = session or requests
    
    try:
        response = cliente_http.post(url_auth, json=credenciais, timeout=10)
        if response.status_code == 200:
            token = response.json()["acesso"]["token"]
            return token
        else:
            st.error(f"❌ Erro ao autenticar: {response.status_code}")
            return None
    except Exception as e:
        st.error(f"❌ Erro de conexão: {str(e)}")
        return None

def _loja_degust_ativa(loja):
    """
    Fallback quando GET /api/loja/loja nao retorna dadosGerais.ativo de forma conclusiva.
    listarLojasFranquia pode trazer situacao / situacaoLoja.
    """
    for key in ("situacaoLoja", "situacao"):
        val = loja.get(key)
        if val is None:
            continue
        s = str(val).strip().upper()
        if s and "INATIV" in s:
            return False
    return True


def _formatar_tabela_de_preco(val):
    """Normaliza tabelaDePreco (string, dict ou None) para exibição."""
    if val is None:
        return None
    if isinstance(val, dict):
        for key in ("descricao", "nome", "nomeTabela", "label", "tabelaDePreco"):
            sub = val.get(key)
            if sub is not None and str(sub).strip():
                return str(sub).strip()
        cod = val.get("codigo") or val.get("id")
        if cod is not None:
            return str(cod).strip()
        return None
    texto = str(val).strip()
    return texto if texto else None


def _extrair_tabela_de_preco_resposta_loja(data):
    """Campo tabelaDePreco em GET /api/loja/loja (raiz ou configuracaoVenda)."""
    if not isinstance(data, dict):
        return None
    configuracao = data.get("configuracaoVenda")
    if not isinstance(configuracao, dict):
        configuracao = {}
    dados_gerais = data.get("dadosGerais")
    if not isinstance(dados_gerais, dict):
        dados_gerais = {}
    for val in (
        data.get("tabelaDePreco"),
        configuracao.get("tabelaDePreco"),
        dados_gerais.get("tabelaDePreco"),
    ):
        texto = _formatar_tabela_de_preco(val)
        if texto:
            return texto
    return None


def _mapa_tabela_preco_por_loja(df_marca, codigos_loja, cliente_http=None, token=None, codfranqueador=None):
    """codigoLoja -> texto da tabela de preço (df_marca e, se faltar, GET /api/loja/loja)."""
    mapa = {}
    if (
        df_marca is not None
        and not df_marca.empty
        and "tabelaDePreco" in df_marca.columns
        and "codigoLoja" in df_marca.columns
    ):
        for _, r in df_marca[["codigoLoja", "tabelaDePreco"]].drop_duplicates(subset=["codigoLoja"]).iterrows():
            try:
                c = int(r["codigoLoja"])
            except (TypeError, ValueError):
                continue
            texto = _formatar_tabela_de_preco(r.get("tabelaDePreco"))
            if texto:
                mapa[c] = texto

    faltantes = [int(c) for c in codigos_loja if int(c) not in mapa]
    if faltantes and cliente_http is not None and token and codfranqueador:

        def _buscar(cod):
            cad = _consultar_cadastro_loja(cliente_http, token, codfranqueador, cod)
            if cad and cad.get("tabelaDePreco"):
                return cod, cad["tabelaDePreco"]
            return cod, None

        workers = max(1, min(8, len(faltantes)))
        with ThreadPoolExecutor(max_workers=workers) as executor:
            for cod, texto in executor.map(_buscar, faltantes):
                if texto:
                    mapa[cod] = texto

    for c in codigos_loja:
        mapa.setdefault(int(c), "N/A")
    return mapa


def _interpretar_campo_ativo_cadastro(val):
    """
    dadosGerais.ativo em LojaResult (GET /api/loja/loja). String no Swagger.
    Retorna True (ativa), False (inativa) ou None (indeterminado).
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    if not s:
        return None
    if s in ("S", "SIM", "1", "T", "TRUE", "Y", "YES", "ATIVO", "ATIVA"):
        return True
    if s in ("N", "NAO", "NÃO", "0", "F", "FALSE", "INATIVO", "INATIVA"):
        return False
    if "INATIV" in s:
        return False
    if "ATIV" in s:
        return True
    return None


def _consultar_cadastro_loja(cliente_http, token, codfranqueador, codigo_loja):
    """GET /api/loja/loja - ativo (dadosGerais) e tabela de preço da loja."""
    url = f"{DEGUST_API_BASE}/api/loja/loja"
    params = {"CodigoFranqueador": int(codfranqueador), "CodigoLoja": int(codigo_loja)}
    headers = {"Authorization": f"Bearer {token}"}
    try:
        r = cliente_http.get(url, params=params, headers=headers, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        if not isinstance(data, dict):
            return None
        dg = data.get("dadosGerais") or {}
        if not isinstance(dg, dict):
            dg = {}
        return {
            "ativo": _interpretar_campo_ativo_cadastro(dg.get("ativo")),
            "tabelaDePreco": _extrair_tabela_de_preco_resposta_loja(data),
        }
    except Exception:
        return None


def _consultar_ativo_cadastro_loja(cliente_http, token, codfranqueador, codigo_loja):
    cadastro = _consultar_cadastro_loja(cliente_http, token, codfranqueador, codigo_loja)
    if cadastro is None:
        return None
    return cadastro.get("ativo")


def _manter_loja_apos_consulta_cadastro(loja, ativo_cadastro):
    if ativo_cadastro is False:
        return False
    if ativo_cadastro is True:
        return True
    return _loja_degust_ativa(loja)


def _filtrar_lojas_por_cadastro_degust(lojas, token, codfranqueador, session_shared=None, max_workers=8):
    """Mantem lojas com dadosGerais.ativo ativa; fallback em situacao da listagem."""
    if not lojas:
        return []

    def processar(loja, http):
        cod = loja.get("codigoLoja")
        try:
            c = int(cod) if cod is not None else None
        except (TypeError, ValueError):
            c = None
        if c is None:
            return _manter_loja_apos_consulta_cadastro(loja, None)
        cadastro = _consultar_cadastro_loja(http, token, codfranqueador, c)
        if cadastro and cadastro.get("tabelaDePreco"):
            loja["tabelaDePreco"] = cadastro["tabelaDePreco"]
        at = cadastro.get("ativo") if cadastro else None
        return _manter_loja_apos_consulta_cadastro(loja, at)

    if session_shared is not None:
        out = []
        for loja in lojas:
            if processar(loja, session_shared):
                out.append(loja)
        return out

    if len(lojas) == 1:
        loja = lojas[0]
        return [loja] if processar(loja, requests) else []

    thread_local = threading.local()

    def http_por_thread():
        if not hasattr(thread_local, "session"):
            thread_local.session = requests.Session()
        return thread_local.session

    workers = max(1, min(max_workers, len(lojas)))

    def work(loja):
        return loja, processar(loja, http_por_thread())

    with ThreadPoolExecutor(max_workers=workers) as executor:
        pairs = list(executor.map(work, lojas))
    return [loja for loja, ok in pairs if ok]


def obter_lojas(token, codfranqueador, session=None):
    """Obtém lista de lojas da franquia com dados completos"""
    url_lojas = f"{DEGUST_API_BASE}/api/loja/listarLojasFranquia?codigoFranquia={codfranqueador}"
    headers = {"Authorization": f"Bearer {token}"}
    cliente_http = session or requests
    
    try:
        response = cliente_http.get(url_lojas, headers=headers, timeout=10)
        if response.status_code == 200:
            lojas = response.json()
            if not isinstance(lojas, list):
                return []
            return _filtrar_lojas_por_cadastro_degust(lojas, token, codfranqueador, session_shared=session)
        else:
            st.error(f"❌ Erro ao buscar lojas: {response.status_code}")
            return []
    except Exception as e:
        st.error(f"❌ Erro de conexão: {str(e)}")
        return []

def consultar_promocoes(token, codfranqueador, lojas_completas, marca, max_workers=8):
    """Consulta promoções de todas as lojas"""
    url_promocoes = f"{DEGUST_API_BASE}/api/produto/consultar-promocoes"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    dados_todas_lojas = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_lojas = len(lojas_completas)

    if total_lojas == 0:
        progress_bar.empty()
        status_text.empty()
        return dados_todas_lojas

    workers = max(1, min(max_workers, total_lojas))
    thread_local = threading.local()

    def _session_thread():
        if not hasattr(thread_local, "session"):
            thread_local.session = requests.Session()
        return thread_local.session

    def _consultar_loja(loja):
        codigo_loja = loja["codigoLoja"]
        nome_loja = loja.get("nomeLoja", "N/A")
        body = {
            "codigoFranquia": codfranqueador,
            "codigoLoja": codigo_loja
        }
        try:
            response = _session_thread().post(url_promocoes, json=body, headers=headers, timeout=10)
            if response.status_code != 200:
                return codigo_loja, nome_loja, []
            dados = response.json() or []
            tabela_preco = loja.get("tabelaDePreco") or "N/A"
            for item in dados:
                item["codigoLoja"] = codigo_loja
                item["nomeLoja"] = nome_loja
                item["marca"] = marca
                item["tabelaDePreco"] = tabela_preco
            return codigo_loja, nome_loja, dados
        except Exception as e:
            return codigo_loja, nome_loja, e

    concluidas = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futuros = [executor.submit(_consultar_loja, loja) for loja in lojas_completas]
        for futuro in as_completed(futuros):
            codigo_loja, nome_loja, resultado = futuro.result()
            concluidas += 1
            status_text.text(f"🔄 Carregando promoções ({concluidas}/{total_lojas}): {nome_loja}")
            progress_bar.progress(concluidas / total_lojas)

            if isinstance(resultado, Exception):
                st.warning(f"⚠️ Erro ao buscar promoções da loja {nome_loja} ({codigo_loja}): {str(resultado)}")
                continue
            if resultado:
                dados_todas_lojas.extend(resultado)
    
    progress_bar.empty()
    status_text.empty()
    
    return dados_todas_lojas

def consultar_produtos_grupo_venda_orientada(token, codfranqueador, lojas_completas, marca, nome_grupo="Promoção"):
    """Consulta produtos por grupo de venda orientada de todas as lojas usando autenticação da API
    
    IMPORTANTE: Utiliza as MESMAS credenciais e token da API de promoções.
    O token é obtido através da função autenticar() que usa CREDENCIAIS["usuario"] e CREDENCIAIS["senha"].
    """
    url = f"{DEGUST_API_BASE}/api/venda-orientada/consultar-produto-por-grupo-venda-orientada"
    
    # Usar o MESMO token e formato de autenticação da API de promoções
    # O token é obtido da mesma função autenticar() que usa as mesmas credenciais
    headers = {
        "Authorization": f"Bearer {token}",  # Mesmo token usado na API de promoções
        "Content-Type": "application/json"
    }
    
    dados_todas_lojas = []
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_lojas = len(lojas_completas)
    
    for idx, loja in enumerate(lojas_completas):
        codigo_loja = loja["codigoLoja"]
        nome_loja = loja.get("nomeLoja", "N/A")
        # Código venda orientada: usar do objeto loja se a API retornar, ou tentar padrão "OUT_2025 EC {NOME_LOJA}"
        cod_venda_orientada_loja = loja.get("codigoVendaOrientada") or loja.get("vendaOrientada") or loja.get("nomeVendaOrientada")
        padroes_venda_orientada = []
        if cod_venda_orientada_loja:
            padroes_venda_orientada.append(str(cod_venda_orientada_loja).strip())
        vo_nome = "OUT_2025 EC " + nome_loja.upper().strip()
        if vo_nome not in padroes_venda_orientada:
            padroes_venda_orientada.append(vo_nome)
        
        status_text.text(f"Carregando produtos do grupo '{nome_grupo}' da loja {idx + 1}/{total_lojas}: {nome_loja}")
        
        nome_grupo_upper = nome_grupo.upper()
        
        if 'BALDES' in nome_grupo_upper:
            bases = [
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "BALDES"},
                {"codigoFranqueador": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "BALDES"},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": nome_grupo},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": nome_grupo.upper()},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "baldes"},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "Baldes"}
            ]
        else:
            bases = [
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "PROMOCAO"},
                {"codigoFranqueador": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "PROMOCAO"},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": nome_grupo},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": nome_grupo.upper()},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "Promoção"},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "PROMOÇÃO"},
                {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "PROMOÇÕES"},
            ]
        # Montar bodies: bases; depois com codigoVendaOrientada, vendaOrientada e nomeVendaOrientada (API pode usar qualquer um)
        bodies_teste = list(bases)
        for padrao in padroes_venda_orientada:
            if not padrao:
                continue
            for b in bases:
                for chave_vo in ("codigoVendaOrientada", "vendaOrientada", "nomeVendaOrientada"):
                    novo = dict(b)
                    novo[chave_vo] = padrao
                    bodies_teste.append(novo)
        
        dados_retornados = None
        for body in bodies_teste:
            try:
                # Usar o MESMO token de autenticação obtido da função autenticar()
                # que utiliza as MESMAS credenciais (CREDENCIAIS["usuario"] e CREDENCIAIS["senha"])
                # usadas na API de promoções
                response = requests.post(url, json=body, headers=headers, timeout=10)
            
                # Verificar se houve erro de autenticação
                if response.status_code == 401:
                    continue  # Tentar próximo body
                elif response.status_code == 403:
                    continue  # Tentar próximo body
                
                if response.status_code == 200:
                    dados = response.json()
                    # Extrair lista de produtos: API pode retornar lista direta ou objeto com chave (data, produtos, itens, etc.)
                    lista_produtos = None
                    if dados is not None:
                        if isinstance(dados, list):
                            lista_produtos = dados if len(dados) > 0 else None
                        elif isinstance(dados, dict):
                            for chave in ("data", "produtos", "itens", "items", "resultado", "content"):
                                if chave in dados and isinstance(dados[chave], list) and len(dados[chave]) > 0:
                                    lista_produtos = dados[chave]
                                    break
                            # Só tratar como objeto único se não for um wrapper com lista vazia
                            if lista_produtos is None and dados and not any(
                                isinstance(dados.get(k), list) for k in ("data", "produtos", "itens", "items", "resultado", "content")
                            ):
                                lista_produtos = [dados]
                    if lista_produtos:
                        dados_retornados = lista_produtos
                        if 'BALDES' in str(body.get('nomeGrupoVendaOrientada', '')).upper():
                            st.info(f"Encontrados {len(dados_retornados)} produtos do grupo 'BALDES' na loja {nome_loja}")
                        break
            except Exception as e:
                continue
        
        # Se POST não retornou dados, tentar GET com query params (algumas APIs aceitam GET)
        if dados_retornados is None:
            try:
                params = {
                    "codigoFranquia": codfranqueador,
                    "codigoLoja": codigo_loja,
                    "nomeGrupoVendaOrientada": "PROMOCAO" if "PROMOCAO" in nome_grupo_upper or "PROMO" in nome_grupo_upper else (nome_grupo if "BALDES" in nome_grupo_upper else "PROMOCAO"),
                }
                if "BALDES" in nome_grupo_upper:
                    params["nomeGrupoVendaOrientada"] = "BALDES"
                resp_get = requests.get(url, params=params, headers=headers, timeout=10)
                if resp_get.status_code == 200:
                    dados = resp_get.json()
                    if isinstance(dados, list) and len(dados) > 0:
                        dados_retornados = dados
                    elif isinstance(dados, dict):
                        for chave in ("data", "produtos", "itens", "items", "resultado", "content"):
                            if chave in dados and isinstance(dados[chave], list) and len(dados[chave]) > 0:
                                dados_retornados = dados[chave]
                                break
            except Exception:
                pass
        
        # Processar dados retornados se encontrou
        if dados_retornados:
            for item in dados_retornados:
                # Normalizar campos da API "Produto por Grupo de Venda Orientada" para o formato do app
                if "codigoProduto" not in item or item.get("codigoProduto") is None:
                    item["codigoProduto"] = (item.get("produto") or item.get("codigoProduto") or item.get("idProduto") or
                                             item.get("id") or item.get("codigo") or "N/A")
                if "descricaoProduto" not in item or item.get("descricaoProduto") is None or item.get("descricaoProduto") == "":
                    item["descricaoProduto"] = (item.get("descricao do produto") or item.get("descricaoProduto") or
                                                item.get("nomeProduto") or item.get("descricao") or "N/A")
                for campo in ("domingo", "segunda", "terca", "quarta", "quinta", "sexta", "sabado", "restricaoHorario", "valorMix", "valorPromocionalMix"):
                    if campo not in item:
                        item[campo] = "N/A"
                item["codigoLoja"] = codigo_loja
                item["nomeLoja"] = nome_loja
                item["marca"] = marca
                # Identificar o grupo de venda orientada baseado no nome_grupo passado como parâmetro
                # ou no nome do grupo retornado pela API
                grupo_identificado = "PROMOCAO"  # Padrão
                nome_grupo_display = "PROMOÇÕES DA UNIDADE"  # Padrão
                
                # Usar o nome_grupo passado como parâmetro (mais confiável)
                nome_grupo_param = str(nome_grupo).upper()
                # Verificar também campos retornados pela API que possam indicar o grupo
                nome_grupo_api = str(item.get('nomeGrupoVendaOrientada', '') or 
                                     item.get('grupoVendaOrientada', '') or 
                                     item.get('nomeGrupo', '') or 
                                     nome_grupo_param).upper()
                
                # Verificar se é "BALDES" - priorizar o parâmetro passado
                if 'BALDES' in nome_grupo_param:
                    grupo_identificado = "BALDES"
                    nome_grupo_display = "BALDES"
                elif 'BALDES' in nome_grupo_api:
                    grupo_identificado = "BALDES"
                    nome_grupo_display = "BALDES"
                elif 'PROMOCAO' in nome_grupo_param or 'PROMOÇÃO' in nome_grupo_param:
                    grupo_identificado = "PROMOCAO"
                    nome_grupo_display = "PROMOÇÕES DA UNIDADE"
                elif 'PROMOCAO' in nome_grupo_api or 'PROMOÇÃO' in nome_grupo_api:
                    grupo_identificado = "PROMOCAO"
                    nome_grupo_display = "PROMOÇÕES DA UNIDADE"
                
                item["grupoVendaOrientada"] = grupo_identificado
                item["nomeGrupo"] = nome_grupo_display
                # Armazenar o nome original do grupo para referência futura
                item["nomeGrupoVendaOrientadaOriginal"] = nome_grupo_api
                item["promocaoAtiva"] = "Sim"
                if "produtoPromocaoAtivo" not in item:
                    item["produtoPromocaoAtivo"] = "Sim"
            dados_todas_lojas.extend(dados_retornados)
        else:
            # Debug temporário: verificar se a API está retornando algo
            # (remover depois de identificar o problema)
            pass
        
        progress_bar.progress((idx + 1) / total_lojas)
    
    progress_bar.empty()
    status_text.empty()
    
    return dados_todas_lojas


def _extrair_lista_resposta(resp):
    """Extrai lista de produtos da resposta da API (lista direta ou dentro de chave)."""
    if resp.status_code != 200:
        return None, 0
    try:
        dados = resp.json()
        if isinstance(dados, list) and len(dados) > 0:
            return dados, len(dados)
        if isinstance(dados, dict):
            for chave in ("data", "produtos", "itens", "items", "resultado", "content"):
                if chave in dados and isinstance(dados[chave], list) and len(dados[chave]) > 0:
                    return dados[chave], len(dados[chave])
    except Exception:
        pass
    return None, 0


def diagnosticar_api_grupo_venda_orientada(marca, codfranqueador, codigo_loja, nome_loja, codigo_venda_orientada=None):
    """
    Faz uma chamada de teste à API consultar-produto-por-grupo-venda-orientada.
    Se codigo_venda_orientada for informado, testa com codigoVendaOrientada, vendaOrientada e nomeVendaOrientada.
    """
    url = f"{DEGUST_API_BASE}/api/venda-orientada/consultar-produto-por-grupo-venda-orientada"
    token = autenticar(codfranqueador)
    if not token:
        return {"erro": "Falha ao autenticar", "status_code": None, "request_body": None, "response_text": None, "tentativas": []}
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    valor_vo = str(codigo_venda_orientada).strip() if codigo_venda_orientada else None

    # Tentar com diferentes nomes do parâmetro de venda orientada
    chaves_vo = ["codigoVendaOrientada", "vendaOrientada", "nomeVendaOrientada"]
    tentativas = []
    body_usado = None
    response_text = "[]"
    resp_len = 0
    status_code = None

    bodies = []
    base = {"codigoFranquia": codfranqueador, "codigoLoja": codigo_loja, "nomeGrupoVendaOrientada": "PROMOCAO"}
    if valor_vo:
        for chave in chaves_vo:
            b = dict(base)
            b[chave] = valor_vo
            bodies.append(b)
    else:
        bodies = [base]

    try:
        for body in bodies:
            r = requests.post(url, json=body, headers=headers, timeout=15)
            lista, n = _extrair_lista_resposta(r)
            tentativas.append({"body": dict(body), "status": r.status_code, "itens": n})
            if n > 0:
                body_usado = body
                status_code = r.status_code
                resp_len = n
                try:
                    response_text = json.dumps(r.json(), indent=2, ensure_ascii=False)[:8000]
                except Exception:
                    response_text = r.text[:8000]
                break
            if body_usado is None:
                body_usado = body
                status_code = r.status_code
                try:
                    response_text = json.dumps(r.json(), indent=2, ensure_ascii=False)[:8000]
                except Exception:
                    response_text = r.text[:8000] if r.text else "[]"
        return {
            "erro": None,
            "status_code": status_code,
            "request_body": body_usado or base,
            "response_text": response_text,
            "response_length": resp_len,
            "tentativas": tentativas,
        }
    except Exception as e:
        return {"erro": str(e), "status_code": None, "request_body": base, "response_text": None, "tentativas": tentativas}


def _ordenar_colunas_download(df):
    """Coloca identificação da loja e tabela de preço no início do Excel."""
    preferidas = [
        "marca", "codigoLoja", "nomeLoja", "tabelaDePreco",
        "nomePromocao", "nomeGrupo", "promocaoAtiva",
        "codigoProduto", "descricaoProduto", "produtoPromocaoAtivo",
    ]
    existentes = [c for c in preferidas if c in df.columns]
    restantes = [c for c in df.columns if c not in existentes]
    return df[existentes + restantes]


def criar_excel_formatado(df):
    """Cria um arquivo Excel formatado a partir de um DataFrame"""
    df = _ordenar_colunas_download(df)
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Promoções"
    
    # Adicionar dados do DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Formatar cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar larguras das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Limitar largura máxima e mínima
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Formatar células de dados (alinhamento)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def analisar_promocoes_por_cobertura(df_marca):
    """Analisa promoções por cobertura de lojas"""
    if df_marca.empty or "nomePromocao" not in df_marca.columns or "codigoLoja" not in df_marca.columns:
        return {}
    
    # Contar quantas lojas cada promoção aparece
    promocao_por_loja = df_marca.groupby("nomePromocao")["codigoLoja"].nunique()
    total_lojas = df_marca["codigoLoja"].nunique()
    
    # Calcular percentual de cobertura
    promocao_cobertura = promocao_por_loja / total_lojas * 100
    
    # Categorizar promoções
    promocoes_100 = promocao_por_loja[promocao_por_loja == total_lojas]
    promocoes_80_plus = promocao_por_loja[promocao_por_loja >= total_lojas * 0.8]
    promocoes_50_plus = promocao_por_loja[promocao_por_loja >= total_lojas * 0.5]
    
    # Top 10 promoções mais comuns
    top_10 = promocao_por_loja.nlargest(10)
    
    # Criar DataFrame com análise
    df_analise = pd.DataFrame({
        'Promoção': promocao_por_loja.index,
        'Lojas': promocao_por_loja.values,
        'Cobertura (%)': promocao_cobertura.values.round(1)
    }).sort_values('Cobertura (%)', ascending=False)
    
    # Adicionar status das promoções
    if "promocaoAtiva" in df_marca.columns:
        status_promocoes = df_marca.groupby("nomePromocao")["promocaoAtiva"].first()
        df_analise['Status'] = df_analise['Promoção'].map(status_promocoes).fillna('N/A')
    
    return {
        'total_lojas': total_lojas,
        'total_promocoes': len(promocao_por_loja),
        'promocoes_100': promocoes_100,
        'promocoes_80_plus': promocoes_80_plus,
        'promocoes_50_plus': promocoes_50_plus,
        'top_10': top_10,
        'df_analise': df_analise,
        'media_cobertura': promocao_cobertura.mean(),
        'mediana_cobertura': promocao_cobertura.median()
    }

@st.cache_data(ttl=300)  # Cache por 5 minutos
def carregar_dados_marca(marca):
    """Carrega dados de uma marca específica, incluindo promoções e produtos do grupo de venda orientada"""
    config = MARCAS_CONFIG[marca]
    codfranqueador = config["codfranqueador"]
    
    # Validar código do franqueador
    if not codfranqueador or codfranqueador == 0:
        st.warning(f"⚠️ Código do franqueador não configurado para {marca}. Por favor, atualize o código em MARCAS_CONFIG.")
        return pd.DataFrame()
    
    with requests.Session() as session:
        # Autenticar
        token = autenticar(codfranqueador, session=session)
        if not token:
            return pd.DataFrame()
        
        # Obter lojas com dados completos
        lojas_completas = obter_lojas(token, codfranqueador, session=session)
        if not lojas_completas:
            return pd.DataFrame()
        
        # Consultar promoções normais (API consultar-promocoes)
        dados = consultar_promocoes(token, codfranqueador, lojas_completas, marca)
    
    # Lógica de listar produtos das categorias "PROMOCAO" e "BALDES" (API grupo venda orientada) foi desativada.
    
    if not dados:
        dados = []
    
    if dados:
        df = pd.DataFrame(dados)
        return df
    else:
        return pd.DataFrame()

def agrupar_por_loja_e_promocao(df):
    """Agrupa dados por loja e depois por promoção, incluindo categorias de grupo de venda orientada"""
    if df.empty:
        return {}
    
    # Primeiro agrupar por loja
    grupos_lojas = {}
    
    # Separar produtos do grupo de venda orientada
    produtos_grupo_venda = []
    produtos_normais = []
    
    # Coletar todas as lojas únicas para garantir que todas tenham a categoria
    lojas_unicas = set()
    
    for _, row in df.iterrows():
        codigo_loja = row.get('codigoLoja', 'N/A')
        nome_loja = row.get('nomeLoja', 'N/A')
        chave_loja = f"{codigo_loja} - {nome_loja}"
        lojas_unicas.add(chave_loja)
        
        # Verificar se é produto do grupo de venda orientada
        # Pode ser "PROMOCAO", "PROMOÇÃO" ou "BALDES"
        grupo_venda = row.get('grupoVendaOrientada')
        if grupo_venda:
            grupo_venda_str = str(grupo_venda).strip().upper()
            # Aceitar "PROMOCAO", "PROMOÇÃO" ou "BALDES"
            if grupo_venda_str in ['PROMOCAO', 'PROMOÇÃO', 'BALDES']:
                produtos_grupo_venda.append(row)
            else:
                produtos_normais.append(row)
        else:
            produtos_normais.append(row)
    
    # Processar produtos normais primeiro
    # Para TODAS as marcas: produtos normais com "PROMOÇÕES" vão para "PROMOÇÕES - {NOME_LOJA}"
    for row in produtos_normais:
        codigo_loja = row.get('codigoLoja', 'N/A')
        nome_loja = row.get('nomeLoja', 'N/A')
        chave_loja = f"{codigo_loja} - {nome_loja}"
        marca_row = row.get('marca', '')
        
        if chave_loja not in grupos_lojas:
            grupos_lojas[chave_loja] = {
                'info_loja': {
                    'codigoLoja': codigo_loja,
                    'nomeLoja': nome_loja,
                    'marca': marca_row,
                    'tabelaDePreco': row.get('tabelaDePreco') or 'N/A',
                },
                'promocoes': {}
            }
        
        # Agrupar promoções dentro da loja
        nome_promocao = row.get('nomePromocao', 'Sem Nome')
        nome_grupo_row = row.get('nomeGrupo', 'N/A')
        sequencia_row = _extrair_sequencia_promocao(row, nome_grupo_row)
        
        # Para TODAS as marcas: se o nome da promoção contém "PROMOÇÕES", garantir que existe
        # e que os produtos normais vão para a seção "PROMOÇÕES - {NOME_LOJA}"
        nome_promocao_normalizado = nome_promocao.upper().replace('Ç', 'C').replace('Õ', 'O').replace('Ã', 'A')
        if 'PROMOCOES' in nome_promocao_normalizado or 'PROMOÇÕES' in nome_promocao_normalizado:
            # Garantir que a promoção "PROMOÇÕES - {NOME_LOJA}" existe
            nome_promocao_base = f"PROMOÇÕES - {nome_loja.upper()}"
            if nome_promocao_base not in grupos_lojas[chave_loja]['promocoes']:
                grupos_lojas[chave_loja]['promocoes'][nome_promocao_base] = {
                    'info_promocao': {
                        'nomePromocao': nome_promocao_base,
                        'promocaoAtiva': 'Sim',
                        'nomeGrupo': 'PROMOÇÕES DA UNIDADE',
                        'sequencia': sequencia_row
                    },
                    'produtos': [],
                    'categorias': {}
                }
            nome_promocao = nome_promocao_base
        
        if nome_promocao not in grupos_lojas[chave_loja]['promocoes']:
            nome_grupo = nome_grupo_row
            sequencia = sequencia_row
            
            grupos_lojas[chave_loja]['promocoes'][nome_promocao] = {
                'info_promocao': {
                    'nomePromocao': nome_promocao,
                    'promocaoAtiva': row.get('promocaoAtiva', 'N/A'),
                    'nomeGrupo': nome_grupo,
                    'sequencia': sequencia
                },
                'produtos': [],
                'categorias': {}  # Nova estrutura para categorias de grupo de venda orientada
            }
        else:
            # Garantir que a estrutura de categorias existe mesmo se a promoção já foi criada
            if 'categorias' not in grupos_lojas[chave_loja]['promocoes'][nome_promocao]:
                grupos_lojas[chave_loja]['promocoes'][nome_promocao]['categorias'] = {}
            # Se a promoção já existe e ainda não tem sequência, preencher quando houver no dado atual.
            info_promocao_existente = grupos_lojas[chave_loja]['promocoes'][nome_promocao]['info_promocao']
            if info_promocao_existente.get('sequencia') in (None, "", "None") and sequencia_row is not None:
                info_promocao_existente['sequencia'] = sequencia_row
        
        # Adicionar produto
        produto = {
            'codigoProduto': row.get('codigoProduto', 'N/A'),
            'descricaoProduto': row.get('descricaoProduto', 'N/A'),
            'produtoPromocaoAtivo': row.get('produtoPromocaoAtivo', 'N/A'),
            'domingo': row.get('domingo', 'N/A'),
            'segunda': row.get('segunda', 'N/A'),
            'terca': row.get('terca', 'N/A'),
            'quarta': row.get('quarta', 'N/A'),
            'quinta': row.get('quinta', 'N/A'),
            'sexta': row.get('sexta', 'N/A'),
            'sabado': row.get('sabado', 'N/A'),
            'restricaoHorario': row.get('restricaoHorario', 'N/A'),
            'autorizaGerente': row.get('autorizaGerente', 'N/A'),
            'taxaServico': row.get('taxaServico', 'N/A'),
            'valorMix': row.get('valorMix', 'N/A'),
            'valorPromocionalMix': row.get('valorPromocionalMix', 'N/A')
        }
        
        grupos_lojas[chave_loja]['promocoes'][nome_promocao]['produtos'].append(produto)
    
    # Garantir que todas as lojas tenham a estrutura de categorias (para compatibilidade)
    # NOTA: Os produtos do grupo de venda orientada agora são adicionados diretamente
    # em 'produtos' da promoção, mas mantemos a estrutura de 'categorias' para não quebrar
    # funcionalidades existentes que possam depender dela
    for chave_loja in lojas_unicas:
        if chave_loja not in grupos_lojas:
            # Se a loja não foi criada ainda, criar estrutura básica
            codigo_loja, nome_loja = chave_loja.split(' - ', 1) if ' - ' in chave_loja else (chave_loja, 'N/A')
            grupos_lojas[chave_loja] = {
                'info_loja': {
                    'codigoLoja': codigo_loja,
                    'nomeLoja': nome_loja,
                    'marca': 'N/A',
                    'tabelaDePreco': 'N/A',
                },
                'promocoes': {}
            }
        
        # Garantir que todas as promoções existentes tenham a estrutura de categorias
        for nome_promocao_existente in grupos_lojas[chave_loja]['promocoes'].keys():
            if 'categorias' not in grupos_lojas[chave_loja]['promocoes'][nome_promocao_existente]:
                grupos_lojas[chave_loja]['promocoes'][nome_promocao_existente]['categorias'] = {}
    
    # Lógica de categorias "PRODUTOS DA CATEGORIA PROMOÇÕES" e "PRODUTOS DA CATEGORIA BALDES" desativada.
    # (Não são mais criadas nem preenchidas; produtos vêm apenas da API consultar-promocoes.)
    
    return grupos_lojas

def _exibir_metrica_texto_completo(label, valor):
    """Rótulo e valor no tamanho do label do st.metric (evita truncar textos longos)."""
    valor_txt = html.escape(str(valor if valor not in (None, "") else "N/A").strip())
    label_txt = html.escape(str(label))
    st.markdown(
        f'<div class="loja-metrica-texto">'
        f'<div class="loja-metrica-label">{label_txt}</div>'
        f'<div class="loja-metrica-valor">{valor_txt}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def exibir_loja_hierarquica(chave_loja, dados_loja, cor_marca, mapa_categoria_vo=None):
    """Exibe uma loja e suas promoções no formato hierárquico"""
    
    mapa_categoria_vo = mapa_categoria_vo or {}
    try:
        cod_loja_int = int(dados_loja["info_loja"]["codigoLoja"])
    except (TypeError, ValueError):
        cod_loja_int = None
    categoria_vo_loja = mapa_categoria_vo.get(cod_loja_int) if cod_loja_int is not None else None
    nome_loja = dados_loja["info_loja"].get("nomeLoja", "N/A")
    # Separar promoções ativas e inativas
    promocoes_ativas = {}
    promocoes_inativas = {}
    
    for nome_promocao, dados_promocao in dados_loja['promocoes'].items():
        if dados_promocao['info_promocao']['promocaoAtiva'] == 'Sim':
            promocoes_ativas[nome_promocao] = dados_promocao
        else:
            promocoes_inativas[nome_promocao] = dados_promocao
    
    # Calcular métricas da loja (incluindo produtos de categorias)
    total_produtos_ativos = 0
    for promocao in promocoes_ativas.values():
        total_produtos_ativos += len(promocao.get('produtos', []))
        for categoria in promocao.get('categorias', {}).values():
            total_produtos_ativos += len(categoria.get('produtos', []))
    
    total_produtos_inativos = 0
    for promocao in promocoes_inativas.values():
        total_produtos_inativos += len(promocao.get('produtos', []))
        for categoria in promocao.get('categorias', {}).values():
            total_produtos_inativos += len(categoria.get('produtos', []))
    
    total_promocoes_ativas = len(promocoes_ativas)
    total_promocoes_inativas = len(promocoes_inativas)
    total_produtos = total_produtos_ativos + total_produtos_inativos
    total_promocoes = total_promocoes_ativas + total_promocoes_inativas
    
    # Criar título do expander com métricas
    tabela_preco = (dados_loja['info_loja'].get('tabelaDePreco') or 'N/A').strip()
    titulo_expander = (
        f"🏪 {chave_loja} | 🟢 {total_promocoes_ativas} Ativas | 🔴 {total_promocoes_inativas} Inativas "
        f"| 🎯 {total_produtos} Produtos | 💰 {tabela_preco}"
    )
    
    # Expander para a loja inteira
    with st.expander(titulo_expander, expanded=False):
        # Informações gerais da loja
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("🟢 Promoções Ativas", total_promocoes_ativas)
        with col2:
            st.metric("🔴 Promoções Inativas", total_promocoes_inativas)
        with col3:
            st.metric("🎯 Total de Produtos", total_produtos)
        with col4:
            st.metric("🏪 Código da Loja", dados_loja['info_loja']['codigoLoja'])
        with col5:
            _exibir_metrica_texto_completo(
                "💰 Tabela de Preço",
                dados_loja['info_loja'].get('tabelaDePreco') or 'N/A',
            )
        
        st.markdown("---")
        
        # Exibir promoções ativas primeiro
        if promocoes_ativas:
            st.markdown("**🟢 Promoções Ativas:**")
            for nome_promocao, dados_promocao in promocoes_ativas.items():
                eh_unidade = _eh_promocao_unidade_por_nome(nome_promocao, nome_loja)
                exibir_promocao_dentro_loja(
                    nome_promocao,
                    dados_promocao,
                    cor_marca,
                    nome_loja=nome_loja,
                    categoria_vo_loja=categoria_vo_loja if eh_unidade else None,
                )
        
        # Exibir promoções inativas consolidadas
        if promocoes_inativas:
            st.markdown("**🔴 Promoções Inativas:**")
            
            # Criar um DataFrame consolidado com todas as promoções inativas
            todas_promocoes_inativas = []
            for nome_promocao, dados_promocao in promocoes_inativas.items():
                for produto in dados_promocao['produtos']:
                    produto_consolidado = {
                        'Promoção': nome_promocao,
                        'Grupo': dados_promocao['info_promocao']['nomeGrupo'],
                        'Status': 'Inativo',
                        **produto
                    }
                    todas_promocoes_inativas.append(produto_consolidado)
            
            if todas_promocoes_inativas:
                df_consolidado = pd.DataFrame(todas_promocoes_inativas)
                
                # Reordenar colunas para melhor visualização
                colunas_ordenadas = [
                    'Promoção', 'Grupo', 'Status',
                    'codigoProduto', 'descricaoProduto',
                    'domingo', 'segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'restricaoHorario',
                    'valorPromocionalMix',
                ]
                
                # Filtrar apenas colunas que existem
                colunas_existentes = [col for col in colunas_ordenadas if col in df_consolidado.columns]
                df_consolidado_ordenado = df_consolidado[colunas_existentes]
                
                # Exibir tabela consolidada
                st.dataframe(
                    df_consolidado_ordenado,
                    use_container_width=True,
                    height=min(400, len(df_consolidado) * 35 + 50)
                )
            else:
                st.info("Nenhuma promoção inativa encontrada.")

def exibir_promocao_dentro_loja(
    nome_promocao,
    dados_promocao,
    cor_marca,
    nome_loja=None,
    categoria_vo_loja=None,
):
    """Exibe uma promoção dentro de uma loja, incluindo categorias de grupo de venda orientada"""
    
    # Container para a promoção
    with st.container():
        # Cabeçalho da promoção
        col1, col2, col3, col4 = st.columns([4, 2, 2, 1])
        
        with col1:
            st.markdown(f"**📦 {nome_promocao}**")
        
        with col2:
            status = "🟢 Ativo" if dados_promocao['info_promocao']['promocaoAtiva'] == 'Sim' else "🔴 Inativo"
            st.markdown(status)
        
        with col3:
            nome_grupo = dados_promocao['info_promocao']['nomeGrupo']
            sequencia = dados_promocao['info_promocao'].get('sequencia')
            if sequencia is not None and _grupo_deve_exibir_sequencia(nome_grupo):
                st.markdown(f"**N{sequencia} - Grupo:** {nome_grupo}")
            else:
                st.markdown(f"**Grupo:** {nome_grupo}")
        
        with col4:
            total_produtos = len(dados_promocao.get('produtos', []))
            total_categorias = len(dados_promocao.get('categorias', {}))
            for categoria in dados_promocao.get('categorias', {}).values():
                total_produtos += len(categoria.get('produtos', []))
            st.markdown(f"**{total_produtos} produtos**")

        eh_promo_unidade = _eh_promocao_unidade_por_nome(nome_promocao, nome_loja)

        # Exibir produtos normais (promoções que não são PROMOÇÕES - LOJA)
        if dados_promocao.get('produtos') and not eh_promo_unidade:
            st.markdown("**📋 Produtos:**")
            df_produtos = pd.DataFrame(dados_promocao['produtos'])
            colunas_existentes = _colunas_tabela_produtos_promocao(df_produtos)
            df_produtos_ordenado = df_produtos[colunas_existentes]
            st.dataframe(
                df_produtos_ordenado,
                use_container_width=True,
                height=min(400, len(df_produtos) * 35 + 50)
            )

        # PROMOÇÕES - LOJA: tabela usa exclusivamente endpoint consultar-promocoes
        if eh_promo_unidade:
            _exibir_tabela_promocoes_loja(nome_loja, dados_promocao.get("produtos"))
            _exibir_categoria_promocao_vo(nome_loja, categoria_vo_loja)
        
        # Exibir categorias de grupo de venda orientada (se houver)
        # Para TODAS as marcas: produtos do grupo de venda orientada aparecem em categoria separada
        categorias = dados_promocao.get('categorias', {})
        if categorias:
            for nome_categoria, dados_categoria in categorias.items():
                produtos_categoria = dados_categoria.get('produtos', [])
                
                # Sempre exibir a categoria, mesmo que esteja vazia
                st.markdown(f"**🏷️ {nome_categoria}**")
                
                if produtos_categoria:
                    # Criar DataFrame dos produtos da categoria
                    df_produtos_categoria = pd.DataFrame(produtos_categoria)
                    nome_grupo = dados_promocao['info_promocao']['nomeGrupo']
                    colunas_existentes = _colunas_tabela_produtos_promocao(df_produtos_categoria)
                    df_produtos_categoria_ordenado = df_produtos_categoria[colunas_existentes]
                    
                    # Exibir tabela de produtos da categoria
                    st.dataframe(
                        df_produtos_categoria_ordenado,
                        use_container_width=True,
                        height=min(400, len(df_produtos_categoria) * 35 + 50)
                    )
                else:
                    # Mostrar mensagem se não houver produtos (mas categoria existe)
                    st.info(f"Nenhum produto encontrado na categoria '{nome_categoria}'.")

        if (
            not dados_promocao.get('produtos')
            and not dados_promocao.get('categorias')
            and not eh_promo_unidade
        ):
            st.info("Nenhum produto encontrado para esta promoção.")
        
        st.markdown("---")

def exibir_promocao_inativa_simples(nome_promocao, dados_promocao, cor_marca):
    """Exibe uma promoção inativa sem expansores aninhados"""
    
    # Container para a promoção inativa
    with st.container():
        # Cabeçalho da promoção inativa
        col1, col2, col3, col4 = st.columns([4, 2, 2, 1])
        
        with col1:
            st.markdown(f"**📦 {nome_promocao}**")
        
        with col2:
            st.markdown("🔴 Inativo")
        
        with col3:
            nome_grupo = dados_promocao['info_promocao']['nomeGrupo']
            sequencia = dados_promocao['info_promocao'].get('sequencia')
            if sequencia is not None and _grupo_deve_exibir_sequencia(nome_grupo):
                st.markdown(f"**N{sequencia} - Grupo:** {nome_grupo}")
            else:
                st.markdown(f"**Grupo:** {nome_grupo}")
        
        with col4:
            st.markdown(f"**{len(dados_promocao['produtos'])} produtos**")
        
        # Exibir produtos diretamente sem expander aninhado
        if dados_promocao['produtos']:
            st.markdown("**📋 Produtos:**")
            # Criar DataFrame dos produtos
            df_produtos = pd.DataFrame(dados_promocao['produtos'])
            colunas_existentes = _colunas_tabela_produtos_promocao(df_produtos)
            df_produtos_ordenado = df_produtos[colunas_existentes]
            
            # Exibir tabela de produtos
            st.dataframe(
                df_produtos_ordenado,
                use_container_width=True,
                height=min(400, len(df_produtos) * 35 + 50)
            )
        else:
            st.info("Nenhum produto encontrado para esta promoção.")
        
        st.markdown("---")

def main():
    # Título
    st.title("⚡ Dashboard de Promoções")
    st.markdown("<div style='text-align: center;'><strong>Marcas Grupo Impettus Degust = Espetto Carioca, Mané, Buteco Seu Rufino & Bendito</strong></div>", unsafe_allow_html=True)
    st.markdown("---")
    
    # Sidebar com filtros e controles
    with st.sidebar:
        st.header("⚙️ Controles")
        
        # Botão de atualizar
        if st.button("🔄 Atualizar Dados", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
        
        # Modo escuro (nativo do Streamlit)
        st.caption("🌙 Modo escuro: use **Settings → Theme → Dark** (canto superior direito).")
        
        st.markdown("---")
        
        # Filtro de marcas
        st.header("🏪 Filtrar por Marca")
        
        filtro_todas = st.checkbox("Todas as Marcas", value=False)
        
        if filtro_todas:
            marcas_selecionadas = list(MARCAS_CONFIG.keys())
        else:
            marcas_selecionadas = []
            for marca in MARCAS_CONFIG.keys():
                if st.checkbox(marca, value=False):
                    marcas_selecionadas.append(marca)
        
        st.markdown("---")
        st.markdown(f"**📅 Última atualização:**  \n{_agora_brasilia_str()}")
    
    # Carregar e exibir dados
    if not marcas_selecionadas:
        st.warning("⚠️ Selecione pelo menos uma marca para visualizar as promoções.")
        return
    
    # Carregar dados de todas as marcas selecionadas
    todos_dados = []
    for marca in marcas_selecionadas:
        with st.spinner(f"Carregando dados de {marca}..."):
            df_marca = carregar_dados_marca(marca)
            if not df_marca.empty:
                todos_dados.append(df_marca)
    
    if not todos_dados:
        st.error("❌ Nenhum dado encontrado para as marcas selecionadas.")
        return
    
    # Combinar todos os dataframes
    df_final = pd.concat(todos_dados, ignore_index=True)
    
    # Botão de download geral (se todas as marcas foram selecionadas)
    if filtro_todas and len(marcas_selecionadas) == len(MARCAS_CONFIG):
        st.markdown("### 📥 Download Geral")
        excel_geral = criar_excel_formatado(df_final)
        st.download_button(
            label="⬇️ Download Todas as Marcas - Todas as Lojas (Excel)",
            data=excel_geral,
            file_name=f"todas_marcas_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_geral"
        )
        st.markdown("---")
    
    # Exibir dados por marca no formato hierárquico
    for marca in marcas_selecionadas:
        df_marca = df_final[df_final["marca"] == marca] if "marca" in df_final.columns else pd.DataFrame()
        
        if not df_marca.empty:
            cor = MARCAS_CONFIG[marca]["cor"]
            
            st.markdown(f"### <span style='color: {cor}'>{marca}</span>", unsafe_allow_html=True)

            codfranqueador = MARCAS_CONFIG[marca]["codfranqueador"]
            with st.spinner(f"Carregando categoria PROMOÇÃO (VO) de {marca}…"):
                n_lojas = df_marca["codigoLoja"].nunique() if "codigoLoja" in df_marca.columns else 0
                mapa_categoria_vo = carregar_mapa_categoria_vo_por_loja(codfranqueador, n_lojas)

            # Agrupar dados por loja e promoção
            grupos_lojas = agrupar_por_loja_e_promocao(df_marca)
            _garantir_secao_promocoes_loja_vo(grupos_lojas, mapa_categoria_vo)
            
            # Informações da marca
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("🏷️ Marcas", df_marca["marca"].nunique() if "marca" in df_marca.columns else 0)
            with col2:
                st.metric("📦 Total de Lojas", len(grupos_lojas))
            with col3:
                st.metric("📊 Total de Produtos", len(df_marca))
            
            _render_bloco_cliques_por_loja(
                marca=marca,
                df_marca=df_marca,
                titulo_expander="📈 AÇÕES PROMOÇÕES DE REDE - Cliques (Vendas) por loja",
                texto_markdown=(
                    "Clique = quantidade vendida do item no relatório de vendas. "
                    "Inclui apenas ações do grupo **PROMOÇÕES DE REDE**."
                ),
                caption_help=(
                    "**Garçom e cliques no período:**\n\n"
                    "-> **Nome do Garçom**: Todos os usuários cadastrados no PDV e que participaram "
                    "na venda da ação.\n\n"
                    "-> **Tabela de preço**: Tabela de preço usada atualmente na loja.\n\n"
                    "-> **Períodos**: Respeitam apenas blocos de até 30 dias (limite da API)."
                ),
                prefixo_key="rede",
                opcoes=listar_opcoes_cliques_promocao(df_marca),
                label_selectbox="Nome da promoção",
                resolver_produtos_fn=lambda nome: resolver_codigos_cliques(
                    df_marca, None, nome
                ),
                origem_label="Promoções de rede",
                mensagem_vazio=(
                    "Não há promoções do grupo PROMOÇÕES REDE para esta marca."
                ),
                excel_prefix="cliques_rede",
            )

            _render_bloco_cliques_por_loja(
                marca=marca,
                df_marca=df_marca,
                titulo_expander="📈 CATEGORIA PROMOÇÃO - Cliques por loja",
                texto_markdown=(
                    "Clique = quantidade vendida do item no relatório de vendas. "
                    "Inclui produtos da **categoria PROMOÇÃO** (venda orientada), "
                    "listados individualmente por nome."
                ),
                caption_help=(
                    "**Garçom e cliques no período:**\n\n"
                    "-> **Nome do Garçom**: Todos os usuários cadastrados no PDV e que participaram "
                    "na venda da ação.\n\n"
                    "-> **Tabela de preço**: Tabela de preço usada atualmente na loja.\n\n"
                    "-> **Períodos**: Respeitam apenas blocos de até 30 dias (limite da API).\n\n"
                    "-> **Lista**: um produto por linha, agregando códigos em todas as lojas da marca."
                ),
                prefixo_key="vo2",
                opcoes=listar_opcoes_cliques_vo(mapa_categoria_vo),
                label_selectbox="CATEGORIA PROMOÇÃO - Cliques por loja",
                resolver_produtos_fn=lambda nome: resolver_codigos_cliques_vo(
                    mapa_categoria_vo, nome
                ),
                origem_label="Categoria PROMOÇÃO (venda orientada)",
                mensagem_vazio=(
                    "Não há produtos na categoria PROMOÇÃO (venda orientada) para esta marca."
                ),
                excel_prefix="cliques_vo2",
            )

            # Análise de Promoções de Rede
            if f'show_modal_{marca}' in st.session_state and st.session_state[f'show_modal_{marca}']:
                with st.expander(f"📊 Promoções de Rede - {marca}", expanded=True):
                    st.markdown(f"### 📊 Análise de Promoções por Cobertura - {marca}")
                    
                    # Realizar análise
                    analise = analisar_promocoes_por_cobertura(df_marca)
                    
                    if analise:
                        # Métricas principais
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("🏪 Total de Lojas", analise['total_lojas'])
                        with col2:
                            st.metric("📦 Total de Promoções", analise['total_promocoes'])
                        with col3:
                            st.metric("📊 Média de Cobertura", f"{analise['media_cobertura']:.1f}%")
                        with col4:
                            st.metric("📈 Mediana de Cobertura", f"{analise['mediana_cobertura']:.1f}%")
                        
                        st.markdown("---")
                        
                        # Promoções por categoria de cobertura
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.markdown("#### 🎯 Promoções em TODAS as Lojas (100%)")
                            if len(analise['promocoes_100']) > 0:
                                for promocao, count in analise['promocoes_100'].items():
                                    st.markdown(f"✅ **{promocao}** ({count} lojas)")
                            else:
                                st.info("Nenhuma promoção aparece em todas as lojas")
                        
                        with col2:
                            st.markdown("#### 🚀 Promoções em 80%+ das Lojas")
                            promocoes_80 = analise['promocoes_80_plus']
                            if len(promocoes_80) > 0:
                                for promocao, count in promocoes_80.items():
                                    percentual = (count / analise['total_lojas']) * 100
                                    st.markdown(f"🟢 **{promocao}** ({count}/{analise['total_lojas']} - {percentual:.1f}%)")
                            else:
                                st.info("Nenhuma promoção com 80%+ de cobertura")
                        
                        with col3:
                            st.markdown("#### 📈 Promoções em 50%+ das Lojas")
                            promocoes_50 = analise['promocoes_50_plus']
                            if len(promocoes_50) > 0:
                                for promocao, count in promocoes_50.items():
                                    percentual = (count / analise['total_lojas']) * 100
                                    st.markdown(f"🟡 **{promocao}** ({count}/{analise['total_lojas']} - {percentual:.1f}%)")
                            else:
                                st.info("Nenhuma promoção com 50%+ de cobertura")
                        
                        st.markdown("---")
                        
                        # Top 10 promoções mais comuns
                        st.markdown("#### 🏆 Top 10 Promoções Mais Comuns")
                        if len(analise['top_10']) > 0:
                            for i, (promocao, count) in enumerate(analise['top_10'].items(), 1):
                                percentual = (count / analise['total_lojas']) * 100
                                st.markdown(f"{i:2d}. **{promocao}** - {count}/{analise['total_lojas']} lojas ({percentual:.1f}%)")
                        
                        st.markdown("---")
                        
                        # Tabela completa
                        st.markdown("#### 📋 Tabela Completa de Promoções")
                        st.dataframe(
                            analise['df_analise'],
                            use_container_width=True,
                            height=400
                        )
                        
                        # Botão para fechar análise
                        if st.button("❌ Fechar Análise", key=f"close_modal_{marca}"):
                            st.session_state[f'show_modal_{marca}'] = False
                            st.rerun()
                    else:
                        st.error("❌ Não foi possível analisar as promoções. Verifique se os dados estão corretos.")
                        if st.button("❌ Fechar Análise", key=f"close_modal_{marca}"):
                            st.session_state[f'show_modal_{marca}'] = False
                            st.rerun()
            
            st.markdown("---")
            
            # Inicializar seleção de lojas para exibição
            lojas_selecionadas_busca = []
            lojas_selecionadas_download = []
            
            # Campo de busca e seletor de lojas para download
            if "nomeLoja" in df_marca.columns and "codigoLoja" in df_marca.columns:
                lojas_info = df_marca[["codigoLoja", "nomeLoja"]].drop_duplicates().sort_values("codigoLoja")
                
                # Criar lista de lojas formatadas para seleção
                lojas_formatadas = []
                for _, loja in lojas_info.iterrows():
                    lojas_formatadas.append(f"{loja['codigoLoja']} - {loja['nomeLoja']}")

                # Seletor de lojas para exibição (substitui busca por digitação)
                st.markdown("**🔍 Buscar Lojas:**")
                lojas_selecionadas_busca = st.multiselect(
                    "Selecione as lojas para exibir (deixe vazio para exibir todas):",
                    options=lojas_formatadas,
                    default=[],
                    key=f"busca_loja_{marca}",
                    help="Seleção de lojas para filtrar a visualização na tela"
                )
                
                # Seletor de lojas para download
                st.markdown("**📥 Selecionar Lojas para Download:**")
                lojas_selecionadas_download = st.multiselect(
                    "Escolha as lojas para download (deixe vazio para todas as lojas da marca):",
                    options=lojas_formatadas,
                    default=[],
                    key=f"lojas_download_{marca}",
                    help="Se nenhuma loja for selecionada, o download incluirá todas as lojas da marca"
                )
                
                # Preparar dados para download baseado na seleção
                df_download = df_marca.copy()
                
                # Se lojas foram selecionadas, filtrar apenas essas lojas
                if lojas_selecionadas_download:
                    # Extrair códigos das lojas selecionadas
                    codigos_selecionados = []
                    for loja_str in lojas_selecionadas_download:
                        codigo = loja_str.split(" - ")[0]
                        try:
                            codigos_selecionados.append(int(codigo))
                        except:
                            codigos_selecionados.append(codigo)
                    
                    # Filtrar DataFrame
                    df_download = df_download[df_download["codigoLoja"].isin(codigos_selecionados)]
                    
                    # Label do botão de download
                    if len(lojas_selecionadas_download) == 1:
                        label_download = f"⬇️ Download {marca} - Loja Selecionada (Excel)"
                    else:
                        label_download = f"⬇️ Download {marca} - {len(lojas_selecionadas_download)} Lojas Selecionadas (Excel)"
                else:
                    # Todas as lojas da marca
                    label_download = f"⬇️ Download {marca} - Todas as Lojas (Excel)"
                
                # Botão de download (Excel formatado)
                excel_file = criar_excel_formatado(df_download)
                st.download_button(
                    label=label_download,
                    data=excel_file,
                    file_name=f"{marca.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_{marca}"
                )
                
                st.markdown("---")
            
            # Exibir lojas e suas promoções no formato hierárquico
            if grupos_lojas:
                st.markdown(f"**📋 Lojas de {marca}:**")
                
                # Filtrar lojas a serem exibidas se houver busca
                grupos_lojas_filtrados = grupos_lojas.copy()
                if lojas_selecionadas_busca and "nomeLoja" in df_marca.columns and "codigoLoja" in df_marca.columns:
                    lojas_escolhidas_set = set(lojas_selecionadas_busca)
                    grupos_lojas_filtrados = {
                        chave: dados for chave, dados in grupos_lojas.items()
                        if chave in lojas_escolhidas_set
                    }
                
                for chave_loja, dados_loja in grupos_lojas_filtrados.items():
                    # Container para cada loja
                    with st.container():
                        exibir_loja_hierarquica(
                            chave_loja,
                            dados_loja,
                            cor,
                            mapa_categoria_vo=mapa_categoria_vo,
                        )
                
                st.markdown("---")
            else:
                st.info(f"Nenhuma loja encontrada para {marca}")

if __name__ == "__main__":
    main()
